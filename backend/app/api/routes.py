import io
import json
import re
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.encoders import jsonable_encoder
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from sqlalchemy import String, and_, case, cast, func, nulls_last, or_
from sqlalchemy.sql import Select
from sqlmodel import Session, select

from ..core.config import get_settings
from ..core.storage import get_remitos_dir_new, resolve_remito_pdf_path
from ..db import get_session
from ..core.security import create_access_token, hash_password, is_legacy_hash, needs_rehash, verify_password
from .deps import SUPERADMIN_EMAIL, get_current_user, require_active_user, require_permissions
from ..models import (
    AuditLog,
    AuditAction,
    Deposit,
    InventoryCount,
    InventoryCountItem,
    InventoryCountStatus,
    Permission,
    Recipe,
    RecipeItem,
    Role,
    RolePermission,
    Order,
    OrderItem,
    Remito,
    RemitoItem,
    Shipment,
    ShipmentItem,
    MermaCause,
    MermaEvent,
    MermaStage,
    MermaType,
    ProductionLine,
    ProductionLot,
    SKU,
    SKUType,
    SemiConversionRule,
    StockLevel,
    StockMovement,
    StockMovementType,
    Supplier,
    PurchaseReceipt,
    PurchaseReceiptItem,
    User,
)
from ..models.common import OrderStatus, RemitoStatus, ShipmentPrepStatus, ShipmentStatus, UnitOfMeasure
from ..schemas import (
    DepositCreate,
    DepositRead,
    DepositUpdate,
    StatusUpdate,
    SKUTypeCreate,
    SKUTypeRead,
    SKUTypeUpdate,
    RecipeCreate,
    RecipeRead,
    RecipeUpdate,
    SKUCreate,
    SKURead,
    SKUUpdate,
    StockMovementCreate,
    StockMovementList,
    StockMovementRead,
    StockMovementTypeCreate,
    StockMovementTypeRead,
    StockMovementTypeUpdate,
    StockLevelRead,
    StockAlertRead,
    StockAlertReport,
    StockReportRead,
    StockSummaryRow,
    MovementSummary,
    UnitRead,
    ProductionLotRead,
    SupplierCreate,
    SupplierRead,
    SupplierUpdate,
    PurchaseReceiptCreate,
    PurchaseReceiptRead,
    PurchaseReceiptItemRead,
    ExpiryReport,
    ExpiryReportRow,
    ExpiryReportStatus,
    LotExpiryUpdate,
    LotExpiryUpdateResponse,
    LoginRequest,
    TokenResponse,
    UserCreate,
    UserRead,
    UserUpdate,
    PermissionRead,
    RolePermissionsUpdate,
    OrderCreate,
    OrderRead,
    OrderSummaryRead,
    OrderUpdate,
    OrderStatusUpdate,
    RemitoRead,
    RemitoItemRead,
    RemitoDispatchRequest,
    RemitoReceiveRequest,
    ShipmentAddOrders,
    ShipmentCreate,
    ShipmentDetail,
    ShipmentItemRead,
    ShipmentItemUpdate,
    ShipmentPrepItemsRequest,
    ShipmentRead,
    ShipmentUpdate,
    MermaCauseCreate,
    MermaCauseRead,
    MermaCauseUpdate,
    MermaEventCreate,
    MermaEventRead,
    MermaTypeCreate,
    MermaTypeRead,
    MermaTypeUpdate,
    ProductionLineCreate,
    ProductionLineRead,
    ProductionLineUpdate,
    InventoryCountCreate,
    InventoryCountRead,
    InventoryCountUpdate,
    InventoryCountItemRead,
    AuditLogRead,
    StockAlertThresholdRead,
    StockAlertThresholdUpdate,
)

public_router = APIRouter()
router = APIRouter(dependencies=[Depends(require_active_user)])
api_router = APIRouter()

AUDIT_EXPORT_LIMIT = 10000

SKU_PRODUCTION_TYPES = {"PT", "SEMI"}
SKU_CONSUMABLE_CODE = "CON"
SKU_SEMI_CODE = "SEMI"
OUTGOING_MOVEMENTS = {"CONSUMPTION", "MERMA", "REMITO"}
INCOMING_MOVEMENTS = {"PRODUCTION", "PURCHASE", "ADJUSTMENT", "TRANSFER"}
LOT_CODE_SEQUENCE_LENGTH = 3

settings = get_settings()


def _encode_changes(payload: dict | list | None) -> dict | None:
    if payload is None:
        return None
    encoded = jsonable_encoder(payload, exclude_none=True)
    return encoded if isinstance(encoded, dict) else {"items": encoded}

def _get_request_ip_address(request: Request | None) -> str | None:
    if not request:
        return None
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    if request.client:
        return request.client.host
    return None


def _build_audit_metadata(request: Request | None) -> dict:
    if not request:
        return {}
    metadata: dict[str, str] = {}
    ip_address = _get_request_ip_address(request)
    if ip_address:
        metadata["ip"] = ip_address
    request_id = request.headers.get("X-Request-ID")
    if request_id:
        metadata["request_id"] = request_id
    user_agent = request.headers.get("User-Agent")
    if user_agent:
        metadata["user_agent"] = user_agent
    return metadata


def _build_audit_context(request: Request | None) -> dict:
    return {
        "ip_address": _get_request_ip_address(request),
        "metadata": _build_audit_metadata(request),
    }


def _user_has_permission(session: Session, user: User, permission_key: str) -> bool:
    if user.email.strip().lower() == SUPERADMIN_EMAIL:
        return True
    if user.role_id is None:
        return False
    result = session.exec(
        select(Permission.key)
        .join(RolePermission, RolePermission.permission_id == Permission.id)
        .where(RolePermission.role_id == user.role_id)
    ).all()
    assigned = {key.lower() for key in result}
    return permission_key.strip().lower() in assigned


def _build_audit_changes(
    event: str,
    entity_type: str,
    entity_id: int | None,
    diff: dict | None = None,
    refs: list[dict] | None = None,
    metadata: dict | None = None,
) -> dict:
    return {
        "event": event,
        "entity": {"type": entity_type, "id": entity_id},
        "diff": diff or {},
        "refs": refs or [],
        "metadata": metadata or {},
    }


def _diff_fields(before: dict, after: dict, fields: list[str]) -> dict:
    changes: dict[str, dict[str, object]] = {}
    for field in fields:
        if before.get(field) != after.get(field):
            changes[field] = {"before": before.get(field), "after": after.get(field)}
    return changes


def _create_diff_from_fields(after: dict, fields: list[str]) -> dict:
    return {field: {"after": after.get(field)} for field in fields}


def _delete_diff_from_fields(before: dict, fields: list[str]) -> dict:
    return {field: {"before": before.get(field)} for field in fields}


def _status_diff(field: str, before: object, after: object) -> dict:
    return {field: {"before": before, "after": after}}


def _build_diff_from_payload(obj: object, update_fields: dict) -> dict:
    changes: dict[str, dict[str, object]] = {}
    for field, new_value in update_fields.items():
        if not hasattr(obj, field):
            continue
        old_value = getattr(obj, field)
        if old_value != new_value:
            changes[field] = {"before": old_value, "after": new_value}
    return changes


def _build_audit_refs(reference_type: str | None, reference_id: int | None) -> list[dict]:
    if not reference_type or reference_id is None:
        return []
    mapping = {
        "REMITO": "remitos",
        "SHIPMENT": "shipments",
        "INVENTORY_COUNT": "inventory_counts",
        "PURCHASE_RECEIPT": "purchase_receipts",
        "ORDER": "orders",
        "MERMA": "mermas",
    }
    reference_key = reference_type.strip().upper()
    entity_type = mapping.get(reference_key, reference_key.lower())
    return [{"type": entity_type, "id": reference_id}]


def _log_audit(
    session: Session,
    entity_type: str,
    entity_id: int | None,
    action: AuditAction,
    user_id: int | None = None,
    changes: dict | list | None = None,
    audit_context: dict | None = None,
) -> None:
    ip_address = None
    metadata = None
    if audit_context:
        ip_address = audit_context.get("ip_address")
        metadata = audit_context.get("metadata")
    encoded_changes = _encode_changes(changes)
    if metadata and isinstance(encoded_changes, dict):
        encoded_changes.setdefault("metadata", metadata)
    session.add(
        AuditLog(
            entity_type=entity_type,
            entity_id=entity_id,
            action=action,
            user_id=user_id,
            changes=encoded_changes,
            ip_address=ip_address,
        )
    )


def _map_user(user: User, session: Session) -> UserRead:
    role_name = None
    if user.role_id:
        role = session.get(Role, user.role_id)
        role_name = role.name if role else None
    return UserRead(
        id=user.id,
        email=user.email,
        full_name=user.full_name,
        role_id=user.role_id,
        role_name=role_name,
        is_active=user.is_active,
    )


def _get_user_role_name(session: Session, user: User) -> str | None:
    if not user.role_id:
        return None
    role = session.get(Role, user.role_id)
    return role.name if role else None


def _normalize_role_name(name: str) -> str:
    return (
        name.strip()
        .upper()
        .replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
        .replace("Ü", "U")
        .replace("Ñ", "N")
    )


def _validate_sku_alert_thresholds(alert_green_min: float | None, alert_yellow_min: float | None) -> None:
    if (alert_green_min is None) != (alert_yellow_min is None):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Las alertas deben configurarse en ambos niveles o en ninguno",
        )


def _is_admin_account(user: User, session: Session) -> bool:
    if user.email.strip().lower() == "admin@local":
        return True
    role_name = _get_user_role_name(session, user)
    if not role_name:
        return False
    return _normalize_role_name(role_name) in {"ADMIN", "ADMINISTRACION"}


def _is_local_account(user: User, session: Session) -> bool:
    role_name = _get_user_role_name(session, user)
    if not role_name:
        return False
    return _normalize_role_name(role_name) in {"ENCARGADO DE LOCALES", "LOCALES", "LOCAL"}


def _get_sku_type_or_404(session: Session, sku_type_id: int) -> SKUType:
    sku_type = session.get(SKUType, sku_type_id)
    if not sku_type:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de SKU no encontrado")
    return sku_type


def _get_semi_units_per_kg(session: Session, sku_id: int) -> float:
    rule = session.exec(select(SemiConversionRule).where(SemiConversionRule.sku_id == sku_id)).first()
    return float(rule.units_per_kg) if rule else 1.0


def _get_production_line_or_404(session: Session, production_line_id: int) -> ProductionLine:
    production_line = session.get(ProductionLine, production_line_id)
    if not production_line:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Línea de producción no encontrada")
    if not production_line.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La línea de producción está inactiva")
    return production_line


def _format_production_date(value: date) -> str:
    return value.strftime("%y%m%d")


def _line_code(line: ProductionLine) -> str:
    return f"L{line.id}"


def _parse_lot_code(lot_code: str) -> tuple[str, str, str, str]:
    parts = lot_code.split("-")
    if len(parts) < 4:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato de lote inválido")
    date_part = parts[0]
    line_part = parts[1]
    seq_part = parts[-1]
    sku_part = "-".join(parts[2:-1])
    return date_part, line_part, sku_part, seq_part


def _validate_lot_code(
    session: Session,
    lot_code: str,
    sku: SKU,
    deposit: Deposit,
    production_line: ProductionLine,
    produced_at: date,
    allow_existing_id: int | None = None,
) -> None:
    date_part, line_part, sku_part, seq_part = _parse_lot_code(lot_code)
    expected_date = _format_production_date(produced_at)
    expected_line = _line_code(production_line)

    if not re.fullmatch(r"\d{6}", date_part):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote debe iniciar con YYMMDD")
    if date_part != expected_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La fecha del lote no coincide con la producción")
    if line_part != expected_line:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote no corresponde a la línea indicada")
    if sku_part != sku.code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote no corresponde al SKU producido")
    if not re.fullmatch(rf"\d{{{LOT_CODE_SEQUENCE_LENGTH}}}", seq_part):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La secuencia del lote debe tener 3 dígitos")

    existing = session.exec(select(ProductionLot).where(ProductionLot.lot_code == lot_code)).first()
    if existing and existing.id != allow_existing_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote ya existe para otro registro")
    if existing:
        if existing.sku_id != sku.id or existing.deposit_id != deposit.id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote pertenece a otro SKU o depósito")
        if existing.production_line_id and existing.production_line_id != production_line.id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote pertenece a otra línea")


def _extract_sequence(lot_code: str) -> int:
    try:
        _, _, _, seq_part = _parse_lot_code(lot_code)
        return int(seq_part)
    except Exception:
        return 0


def _generate_lot_code(session: Session, sku: SKU, production_line: ProductionLine, produced_at: date) -> str:
    prefix = f"{_format_production_date(produced_at)}-{_line_code(production_line)}-{sku.code}"
    existing_codes = session.exec(
        select(ProductionLot.lot_code).where(
            ProductionLot.sku_id == sku.id,
            ProductionLot.production_line_id == production_line.id,
            ProductionLot.produced_at == produced_at,
        )
    ).all()
    max_seq = max((_extract_sequence(code) for code in existing_codes), default=0)
    return f"{prefix}-{max_seq + 1:0{LOT_CODE_SEQUENCE_LENGTH}d}"


def _upsert_semi_conversion_rule(session: Session, sku_id: int, units_per_kg: float) -> None:
    if units_per_kg <= 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La conversión de SEMI debe ser mayor a cero")
    rule = session.exec(select(SemiConversionRule).where(SemiConversionRule.sku_id == sku_id)).first()
    if rule:
        rule.units_per_kg = units_per_kg
        rule.updated_at = datetime.utcnow()
    else:
        rule = SemiConversionRule(sku_id=sku_id, units_per_kg=units_per_kg)
    session.add(rule)


def _delete_semi_conversion_rule(session: Session, sku_id: int) -> None:
    existing = session.exec(select(SemiConversionRule).where(SemiConversionRule.sku_id == sku_id)).first()
    if existing:
        session.delete(existing)


def _convert_to_base_quantity(sku: SKU, quantity: float, unit: UnitOfMeasure | None, session: Session) -> float:
    session.refresh(sku, attribute_names=["sku_type"])
    if sku.sku_type and sku.sku_type.code == SKU_SEMI_CODE:
        units_per_kg = _get_semi_units_per_kg(session, sku.id)
        if unit in (None, UnitOfMeasure.KG):
            return quantity
        if unit == UnitOfMeasure.UNIT:
            return quantity / units_per_kg
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unidad no soportada para SEMI")
    return quantity


def _get_recipe_for_product(session: Session, product_id: int) -> Recipe | None:
    return session.exec(select(Recipe).where(Recipe.product_id == product_id, Recipe.is_active.is_(True))).first()


def _calculate_consumption_by_lot(
    session: Session, component_id: int, deposit_id: int, required_quantity: float, strict: bool = False
) -> list[tuple[ProductionLot | None, float]]:
    lots = session.exec(
        select(ProductionLot)
        .where(
            ProductionLot.sku_id == component_id,
            ProductionLot.deposit_id == deposit_id,
            ProductionLot.is_blocked.is_(False),
        )
        .order_by(nulls_last(ProductionLot.expiry_date), ProductionLot.produced_at, ProductionLot.id)
    ).all()

    remaining = required_quantity
    consumptions: list[tuple[ProductionLot | None, float]] = []

    for lot in lots:
        if remaining <= 0:
            break
        available = float(lot.remaining_quantity)
        if available <= 0:
            continue
        take = min(remaining, available)
        consumptions.append((lot, take))
        remaining -= take

    if remaining > 0:
        if strict:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Stock insuficiente para cubrir la cantidad requerida",
            )
        if lots:
            target_lot = consumptions[-1][0] if consumptions else lots[-1]
            if consumptions and consumptions[-1][0] == target_lot:
                consumptions[-1] = (target_lot, consumptions[-1][1] + remaining)
            else:
                consumptions.append((target_lot, remaining))
            remaining = 0
        else:
            consumptions.append((None, remaining))

    return consumptions


def _consume_recipe_components(
    session: Session,
    product: SKU,
    deposit: Deposit,
    production_lot: ProductionLot | None,
    produced_quantity: float,
    reference: str | None,
    movement_date: date,
    created_by_user_id: int | None,
) -> None:
    if produced_quantity <= 0:
        return

    recipe = _get_recipe_for_product(session, product.id)
    if not recipe:
        return
    session.refresh(recipe, attribute_names=["items"])
    if not recipe.items:
        return

    consumption_type = _get_movement_type_by_code(session, "CONSUMPTION")
    reference_value = reference or (production_lot.lot_code if production_lot else f"PROD-{product.code}")

    for item in recipe.items:
        component = session.get(SKU, item.component_id)
        if not component:
            continue

        required_quantity = produced_quantity * item.quantity
        consumptions = _calculate_consumption_by_lot(session, component.id, deposit.id, required_quantity)

        for lot, quantity in consumptions:
            movement_payload = StockMovementCreate(
                sku_id=component.id,
                deposit_id=deposit.id,
                movement_type_id=consumption_type.id,
                quantity=quantity,
                reference=reference_value,
                lot_code=lot.lot_code if lot else None,
                production_lot_id=lot.id if lot else None,
                movement_date=movement_date,
                created_by_user_id=created_by_user_id,
            )
            _apply_stock_movement(session, movement_payload, allow_negative_balance=True, audit_movement=True)


def _map_sku(sku: SKU, session: Session) -> SKURead:
    session.refresh(sku, attribute_names=["sku_type"])
    type_code = sku.sku_type.code if sku.sku_type else ""
    type_label = sku.sku_type.label if sku.sku_type else ""
    secondary_unit = UnitOfMeasure.UNIT if type_code == SKU_SEMI_CODE else None
    units_per_kg = _get_semi_units_per_kg(session, sku.id) if type_code == SKU_SEMI_CODE else None
    return SKURead(
        id=sku.id,
        code=sku.code,
        name=sku.name,
        sku_type_id=sku.sku_type_id,
        sku_type_code=type_code,
        sku_type_label=type_label,
        unit=sku.unit,
        secondary_unit=secondary_unit,
        units_per_kg=units_per_kg,
        notes=sku.notes,
        is_active=sku.is_active,
        alert_green_min=sku.alert_green_min,
        alert_yellow_min=sku.alert_yellow_min,
    )


def _has_alert_thresholds(sku: SKU) -> bool:
    return sku.alert_green_min is not None and sku.alert_yellow_min is not None


def _get_stock_alert_status(quantity: float, sku: SKU) -> str:
    if sku.alert_green_min is not None and sku.alert_yellow_min is not None and quantity >= sku.alert_green_min:
        return "green"
    if sku.alert_green_min is not None and sku.alert_yellow_min is not None and quantity >= sku.alert_yellow_min:
        return "yellow"
    if sku.alert_green_min is not None and sku.alert_yellow_min is not None:
        return "red"
    return "none"


EXPIRY_RED_DAYS = 7
EXPIRY_YELLOW_DAYS = 15


def _get_expiry_status(expiry_date: date | None, today: date) -> tuple[ExpiryReportStatus, int | None]:
    if not expiry_date:
        return ExpiryReportStatus.NONE, None
    days = (expiry_date - today).days
    if days <= EXPIRY_RED_DAYS:
        return ExpiryReportStatus.RED, days
    if days <= EXPIRY_YELLOW_DAYS:
        return ExpiryReportStatus.YELLOW, days
    return ExpiryReportStatus.GREEN, days


def _map_stock_level(level: StockLevel, session: Session) -> StockLevelRead:
    session.refresh(level, attribute_names=["sku", "deposit"])
    return StockLevelRead(
        deposit_id=level.deposit_id,
        deposit_name=level.deposit.name,
        sku_id=level.sku_id,
        sku_code=level.sku.code,
        sku_name=level.sku.name,
        quantity=level.quantity,
        alert_status=_get_stock_alert_status(level.quantity, level.sku),
        alert_green_min=level.sku.alert_green_min,
        alert_yellow_min=level.sku.alert_yellow_min,
    )


def _map_production_lot(lot: ProductionLot, session: Session) -> ProductionLotRead:
    session.refresh(lot, attribute_names=["sku", "deposit", "production_line"])
    return ProductionLotRead(
        id=lot.id,
        lot_code=lot.lot_code,
        sku_id=lot.sku_id,
        sku_code=lot.sku.code if lot.sku else "",
        sku_name=lot.sku.name if lot.sku else "",
        deposit_id=lot.deposit_id,
        deposit_name=lot.deposit.name if lot.deposit else "",
        production_line_id=lot.production_line_id,
        production_line_name=lot.production_line.name if lot.production_line else None,
        produced_quantity=float(lot.produced_quantity),
        remaining_quantity=float(lot.remaining_quantity),
        produced_at=lot.produced_at,
        expiry_date=lot.expiry_date,
        is_blocked=lot.is_blocked,
        notes=lot.notes,
    )


def _get_movement_type_or_404(session: Session, movement_type_id: int) -> StockMovementType:
    movement_type = session.get(StockMovementType, movement_type_id)
    if not movement_type:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de movimiento no encontrado")
    return movement_type


def _get_movement_type_by_code(session: Session, code: str) -> StockMovementType:
    movement_type = session.exec(
        select(StockMovementType).where(StockMovementType.code == code.strip().upper())
    ).first()
    if not movement_type:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de movimiento no configurado")
    return movement_type


def _ensure_store_destination(session: Session, destination_id: int) -> Deposit:
    deposit = session.get(Deposit, destination_id)
    if not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Destino no encontrado")
    if not deposit.is_store:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El destino debe ser un local definido")
    return deposit


def _is_integer_value(value: object) -> bool:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    if isinstance(value, float) and not value.is_integer():
        return False
    return True


def _validate_order_items(session: Session, items: list[dict]) -> None:
    sku_ids = {item["sku_id"] for item in items}
    skus = session.exec(select(SKU).where(SKU.id.in_(sku_ids))).all()
    if len(skus) != len(sku_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún SKU no existe")

    sku_map = {sku.id: sku for sku in skus}
    for item in items:
        quantity = item.get("quantity")
        if quantity is None:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Cantidad requerida")
        if not _is_integer_value(quantity):
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="La cantidad debe ser un número entero")
        if quantity <= 0:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="La cantidad debe ser mayor a cero")

        current_stock = item.get("current_stock")
        if current_stock is None:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Stock actual requerido")
        if not _is_integer_value(current_stock):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="El stock actual debe ser un número entero",
            )
        if current_stock < 0:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="El stock actual no puede ser negativo")

        sku = sku_map.get(item["sku_id"])
        if not sku:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU no encontrado")
        session.refresh(sku, attribute_names=["sku_type"])
        type_code = sku.sku_type.code if sku.sku_type else ""
        if not (sku.sku_type and sku.sku_type.is_active):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún tipo de SKU está inactivo")
        if type_code in {"MP", "SEMI"}:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se permiten MP o SEMI en pedidos",
            )
        if not sku.is_active:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún SKU está inactivo")


def _ensure_order_transition(order: Order, next_status: OrderStatus) -> None:
    if next_status == OrderStatus.DRAFT and order.status != OrderStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede volver a borrador")
    if next_status == OrderStatus.SUBMITTED and order.status != OrderStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo un borrador puede enviarse")
    if next_status in {
        OrderStatus.PREPARED,
        OrderStatus.PARTIALLY_PREPARED,
        OrderStatus.PARTIALLY_DISPATCHED,
        OrderStatus.DISPATCHED,
    }:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El estado de despacho solo se actualiza desde envíos",
        )
    if next_status == OrderStatus.CANCELLED and order.status not in {
        OrderStatus.DRAFT,
        OrderStatus.SUBMITTED,
    }:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El pedido no puede cancelarse en este estado")


def _validate_required_delivery_date(required_delivery_date: date | None) -> None:
    if required_delivery_date is None:
        return
    today = date.today()
    max_date = today + timedelta(days=60)
    if required_delivery_date < today or required_delivery_date > max_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La fecha requerida debe estar entre hoy y los próximos 60 días",
        )


def _serialize_order_item(item: OrderItem) -> dict:
    return {
        "sku_id": item.sku_id,
        "quantity": float(item.quantity),
        "current_stock": item.current_stock,
    }


def _serialize_order_item_payload(item: dict) -> dict:
    return {
        "sku_id": item["sku_id"],
        "quantity": float(item["quantity"]),
        "current_stock": item.get("current_stock"),
    }


def _get_dispatched_quantities(session: Session, order_item_ids: list[int]) -> dict[int, float]:
    if not order_item_ids:
        return {}
    statement = (
        select(ShipmentItem.order_item_id, func.sum(ShipmentItem.quantity))
        .join(Shipment)
        .where(
            ShipmentItem.order_item_id.in_(order_item_ids),
            Shipment.status == ShipmentStatus.DISPATCHED,
        )
        .group_by(ShipmentItem.order_item_id)
    )
    rows = session.exec(statement).all()
    return {row[0]: float(row[1] or 0) for row in rows}


def _get_prepared_quantities(
    session: Session,
    order_item_ids: list[int],
    exclude_shipment_id: int | None = None,
) -> dict[int, float]:
    if not order_item_ids:
        return {}
    statement = (
        select(
            ShipmentItem.order_item_id,
            func.coalesce(func.sum(case((ShipmentItem.is_ready == True, ShipmentItem.quantity), else_=0)), 0),  # noqa: E712
        )
        .join(Shipment)
        .where(
            ShipmentItem.order_item_id.in_(order_item_ids),
            Shipment.status.in_({ShipmentStatus.DRAFT, ShipmentStatus.CONFIRMED}),
        )
    )
    if exclude_shipment_id is not None:
        statement = statement.where(ShipmentItem.shipment_id != exclude_shipment_id)
    statement = statement.group_by(ShipmentItem.order_item_id)
    rows = session.exec(statement).all()
    return {row[0]: float(row[1] or 0) for row in rows}


def _get_assigned_shipment_quantities(
    session: Session,
    order_item_ids: list[int],
    exclude_shipment_id: int | None = None,
) -> dict[int, float]:
    if not order_item_ids:
        return {}
    statement = (
        select(ShipmentItem.order_item_id, func.sum(ShipmentItem.quantity))
        .join(Shipment)
        .where(
            ShipmentItem.order_item_id.in_(order_item_ids),
            Shipment.status.in_({ShipmentStatus.DRAFT, ShipmentStatus.CONFIRMED}),
        )
    )
    if exclude_shipment_id is not None:
        statement = statement.where(ShipmentItem.shipment_id != exclude_shipment_id)
    statement = statement.group_by(ShipmentItem.order_item_id)
    rows = session.exec(statement).all()
    return {row[0]: float(row[1] or 0) for row in rows}


def _get_assigned_quantities(
    session: Session,
    order_item_ids: list[int],
    exclude_shipment_id: int | None = None,
) -> dict[int, float]:
    prepared = _get_assigned_shipment_quantities(session, order_item_ids, exclude_shipment_id=exclude_shipment_id)
    dispatched = _get_dispatched_quantities(session, order_item_ids)
    combined = prepared.copy()
    for order_item_id, quantity in dispatched.items():
        combined[order_item_id] = combined.get(order_item_id, 0.0) + quantity
    return combined


def _sync_order_statuses(
    session: Session,
    order_ids: set[int],
    current_user: User,
    shipment_id: int | None = None,
    audit_context: dict | None = None,
) -> None:
    if not order_ids:
        return
    orders = session.exec(select(Order).where(Order.id.in_(order_ids))).all()
    for order in orders:
        order_items = session.exec(select(OrderItem).where(OrderItem.order_id == order.id)).all()
        order_item_ids = [item.id for item in order_items]
        dispatched_quantities = _get_dispatched_quantities(session, order_item_ids)
        prepared_quantities = _get_prepared_quantities(session, order_item_ids)
        all_dispatched = True
        any_dispatched = False
        all_prepared = True
        any_prepared = False
        for order_item in order_items:
            dispatched = dispatched_quantities.get(order_item.id, 0.0)
            prepared = prepared_quantities.get(order_item.id, 0.0)
            if dispatched <= 0:
                all_dispatched = False
            if dispatched > 0:
                any_dispatched = True
            if dispatched < float(order_item.quantity):
                all_dispatched = False
            if prepared <= 0:
                all_prepared = False
            if prepared > 0:
                any_prepared = True
            if prepared < float(order_item.quantity):
                all_prepared = False
        if any_dispatched:
            new_status = OrderStatus.DISPATCHED if all_dispatched else OrderStatus.PARTIALLY_DISPATCHED
        elif any_prepared:
            new_status = OrderStatus.PREPARED if all_prepared else OrderStatus.PARTIALLY_PREPARED
        else:
            new_status = OrderStatus.SUBMITTED if order.status != OrderStatus.DRAFT else order.status
        if order.status != new_status:
            before_status = order.status
            order.status = new_status
            order.updated_at = datetime.utcnow()
            order.updated_by_user_id = current_user.id
            session.add(order)
            _log_audit(
                session,
                "orders",
                order.id,
                AuditAction.STATUS,
                current_user.id,
                _build_audit_changes(
                    "status",
                    "orders",
                    order.id,
                    diff=_status_diff("status", before_status, new_status),
                    refs=_build_audit_refs("SHIPMENT", shipment_id) if shipment_id else [],
                    metadata=(audit_context or {}).get("metadata", {}),
                ),
                audit_context=audit_context,
            )


def _get_latest_shipment_date(session: Session, order_id: int) -> date | None:
    statement = (
        select(Shipment.estimated_delivery_date)
        .join(ShipmentItem, ShipmentItem.shipment_id == Shipment.id)
        .where(
            ShipmentItem.order_id == order_id,
            Shipment.status.in_({ShipmentStatus.CONFIRMED, ShipmentStatus.DISPATCHED}),
        )
        .order_by(Shipment.estimated_delivery_date.desc(), Shipment.id.desc())
    )
    return session.exec(statement).first()


def _order_locked_by_shipment(session: Session, order_id: int) -> bool:
    statement = (
        select(ShipmentItem.id)
        .join(Shipment)
        .where(
            ShipmentItem.order_id == order_id,
            Shipment.status.in_({ShipmentStatus.CONFIRMED, ShipmentStatus.DISPATCHED}),
        )
        .limit(1)
    )
    return session.exec(statement).first() is not None


def _order_has_shipment_assignments(session: Session, order_id: int) -> bool:
    statement = select(ShipmentItem.id).where(ShipmentItem.order_id == order_id).limit(1)
    return session.exec(statement).first() is not None


def _order_has_ready_picking(session: Session, order_id: int) -> bool:
    statement = (
        select(ShipmentItem.id)
        .join(Shipment)
        .where(
            ShipmentItem.order_id == order_id,
            ShipmentItem.is_ready.is_(True),
            Shipment.status.in_({ShipmentStatus.DRAFT, ShipmentStatus.CONFIRMED, ShipmentStatus.DISPATCHED}),
        )
        .limit(1)
    )
    return session.exec(statement).first() is not None


def _remove_order_from_draft_shipments(session: Session, order_id: int) -> None:
    items = session.exec(
        select(ShipmentItem)
        .join(Shipment)
        .where(ShipmentItem.order_id == order_id, Shipment.status == ShipmentStatus.DRAFT)
    ).all()
    for item in items:
        session.delete(item)


def _ensure_order_cancel_allowed(session: Session, order: Order) -> None:
    if order.status == OrderStatus.DRAFT:
        return
    if order.status == OrderStatus.SUBMITTED:
        if _order_has_ready_picking(session, order.id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se puede cancelar: el pedido ya tiene picking iniciado.",
            )
        if _order_locked_by_shipment(session, order.id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se puede cancelar: el pedido está en un envío confirmado o despachado.",
            )
        return


def _map_order(order: Order, session: Session) -> OrderRead:
    session.refresh(order, attribute_names=["items"])
    order_item_ids = [item.id for item in order.items]
    dispatched_quantities = _get_dispatched_quantities(session, order_item_ids)
    prepared_quantities = _get_prepared_quantities(session, order_item_ids)
    items = []
    for item in order.items:
        sku = session.get(SKU, item.sku_id)
        sku_code = sku.code if sku else str(item.sku_id)
        sku_name = sku.name if sku else f"SKU {item.sku_id}"
        quantity_value = float(item.quantity)
        has_legacy_decimal = not quantity_value.is_integer()
        dispatched_quantity = dispatched_quantities.get(item.id, 0.0)
        prepared_quantity = prepared_quantities.get(item.id, 0.0)
        pending_quantity = max(quantity_value - prepared_quantity - dispatched_quantity, 0.0)
        items.append(
            {
                "id": item.id,
                "sku_id": item.sku_id,
                "sku_code": sku_code,
                "sku_name": sku_name,
                "quantity": quantity_value,
                "current_stock": item.current_stock,
                "prepared_quantity": prepared_quantity,
                "dispatched_quantity": dispatched_quantity,
                "pending_quantity": pending_quantity,
                "has_legacy_decimal": has_legacy_decimal,
                "quantity_raw": quantity_value if has_legacy_decimal else None,
            }
        )

    created_by_name = None
    updated_by_name = None
    if order.created_by_user_id:
        user = session.get(User, order.created_by_user_id)
        created_by_name = user.full_name if user else None
    if order.updated_by_user_id:
        user = session.get(User, order.updated_by_user_id)
        updated_by_name = user.full_name if user else None

    return OrderRead(
        id=order.id,
        destination=order.destination,
        destination_deposit_id=order.destination_deposit_id,
        requested_for=order.requested_for,
        required_delivery_date=order.required_delivery_date,
        requested_by=order.requested_by,
        estimated_delivery_date=_get_latest_shipment_date(session, order.id),
        status=order.status,
        notes=order.notes,
        plant_internal_note=order.plant_internal_note,
        created_at=order.created_at,
        cancelled_at=order.cancelled_at,
        cancelled_by_user_id=order.cancelled_by_user_id,
        cancelled_by_name=order.cancelled_by_name,
        created_by_user_id=order.created_by_user_id,
        created_by_name=created_by_name,
        updated_by_user_id=order.updated_by_user_id,
        updated_by_name=updated_by_name,
        items=items,
    )


def _get_deposit_or_404(session: Session, deposit_id: int) -> Deposit:
    deposit = session.get(Deposit, deposit_id)
    if not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Depósito no encontrado")
    return deposit


def _default_source_deposit(session: Session) -> Deposit:
    deposit = session.exec(select(Deposit).where(Deposit.is_store.is_(False))).first()
    if not deposit:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay depósitos configurados como origen")
    return deposit


def _get_remito_or_404(session: Session, remito_id: int) -> Remito:
    remito = session.get(Remito, remito_id)
    if not remito:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Remito no encontrado")
    return remito


def _get_shipment_or_404(session: Session, shipment_id: int) -> Shipment:
    shipment = session.get(Shipment, shipment_id)
    if not shipment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Envío no encontrado")
    return shipment


def _map_order_summary(order: Order) -> OrderSummaryRead:
    return OrderSummaryRead(
        id=order.id,
        status=order.status,
        destination=order.destination,
        requested_for=order.requested_for,
        required_delivery_date=order.required_delivery_date,
    )


def _get_shipment_orders(session: Session, shipment_id: int) -> list[OrderSummaryRead]:
    statement = (
        select(Order)
        .join(ShipmentItem, ShipmentItem.order_id == Order.id)
        .where(ShipmentItem.shipment_id == shipment_id)
        .order_by(Order.id)
    )
    orders = session.exec(statement).all()
    seen: set[int] = set()
    summaries: list[OrderSummaryRead] = []
    for order in orders:
        if order.id in seen:
            continue
        summaries.append(_map_order_summary(order))
        seen.add(order.id)
    return summaries


def _map_remito(remito: Remito, session: Session) -> RemitoRead:
    session.refresh(remito, attribute_names=["items", "source_deposit", "destination_deposit"])
    items: list[RemitoItemRead] = []
    for item in remito.items:
        sku = session.get(SKU, item.sku_id)
        items.append(
            RemitoItemRead(
                id=item.id,
                remito_id=item.remito_id,
                sku_id=item.sku_id,
                sku_code=sku.code if sku else str(item.sku_id),
                sku_name=sku.name if sku else f"SKU {item.sku_id}",
                quantity=item.quantity,
                lot_code=item.lot_code,
            )
        )
    created_by_name = None
    updated_by_name = None
    if remito.created_by_user_id:
        user = session.get(User, remito.created_by_user_id)
        created_by_name = user.full_name if user else None
    if remito.updated_by_user_id:
        user = session.get(User, remito.updated_by_user_id)
        updated_by_name = user.full_name if user else None
    origin_orders: list[OrderSummaryRead] = []
    if remito.shipment_id:
        origin_orders = _get_shipment_orders(session, remito.shipment_id)
    elif remito.order_id:
        order = session.get(Order, remito.order_id)
        if order:
            origin_orders = [_map_order_summary(order)]
    return RemitoRead(
        id=remito.id,
        order_id=remito.order_id,
        shipment_id=remito.shipment_id,
        status=remito.status,
        destination=remito.destination,
        source_deposit_id=remito.source_deposit_id,
        destination_deposit_id=remito.destination_deposit_id,
        source_deposit_name=remito.source_deposit.name if remito.source_deposit else None,
        destination_deposit_name=remito.destination_deposit.name if remito.destination_deposit else None,
        issue_date=remito.issue_date,
        dispatched_at=remito.dispatched_at,
        received_at=remito.received_at,
        cancelled_at=remito.cancelled_at,
        created_at=remito.created_at,
        pdf_path=remito.pdf_path,
        created_by_user_id=remito.created_by_user_id,
        created_by_name=created_by_name,
        updated_by_user_id=remito.updated_by_user_id,
        updated_by_name=updated_by_name,
        origin_orders=origin_orders,
        items=items,
    )


def _map_shipment_item(item: ShipmentItem, session: Session, dispatched_quantities: dict[int, float]) -> ShipmentItemRead:
    order_item = session.get(OrderItem, item.order_item_id)
    sku = session.get(SKU, order_item.sku_id) if order_item else None
    ordered_quantity = float(order_item.quantity) if order_item else 0.0
    dispatched_quantity = dispatched_quantities.get(item.order_item_id, 0.0)
    remaining_quantity = max(ordered_quantity - dispatched_quantity, 0.0)
    return ShipmentItemRead(
        id=item.id,
        shipment_id=item.shipment_id,
        order_id=item.order_id,
        order_item_id=item.order_item_id,
        sku_id=order_item.sku_id if order_item else 0,
        sku_code=sku.code if sku else str(order_item.sku_id) if order_item else "",
        sku_name=sku.name if sku else f"SKU {order_item.sku_id}" if order_item else "",
        quantity=item.quantity,
        ordered_quantity=ordered_quantity,
        dispatched_quantity=dispatched_quantity,
        remaining_quantity=remaining_quantity,
        is_ready=item.is_ready,
    )


def _calculate_prep_status(total_count: int, ready_count: int) -> ShipmentPrepStatus:
    if total_count <= 0 or ready_count == 0:
        return ShipmentPrepStatus.PENDING
    if ready_count >= total_count:
        return ShipmentPrepStatus.READY
    return ShipmentPrepStatus.PARTIAL


def _get_prep_counts(session: Session, shipment_id: int) -> tuple[int, int]:
    statement = select(
        func.count(ShipmentItem.id),
        func.coalesce(func.sum(case((ShipmentItem.is_ready == True, 1), else_=0)), 0),  # noqa: E712
    ).where(ShipmentItem.shipment_id == shipment_id)
    total_count, ready_count = session.exec(statement).one()
    return int(total_count or 0), int(ready_count or 0)


def _map_shipment(shipment: Shipment, session: Session, include_items: bool = False) -> ShipmentRead | ShipmentDetail:
    session.refresh(shipment, attribute_names=["items", "deposit"])
    if include_items:
        total_count = len(shipment.items)
        ready_count = sum(1 for item in shipment.items if item.is_ready)
    else:
        total_count, ready_count = _get_prep_counts(session, shipment.id)
    base = ShipmentRead(
        id=shipment.id,
        deposit_id=shipment.deposit_id,
        deposit_name=shipment.deposit.name if shipment.deposit else None,
        estimated_delivery_date=shipment.estimated_delivery_date,
        status=shipment.status,
        prep_status=_calculate_prep_status(total_count, ready_count),
        created_at=shipment.created_at,
        updated_at=shipment.updated_at,
    )
    if not include_items:
        return base
    dispatched_quantities = _get_dispatched_quantities(session, [item.order_item_id for item in shipment.items])
    items = [_map_shipment_item(item, session, dispatched_quantities) for item in shipment.items]
    orders = _get_shipment_orders(session, shipment.id)
    return ShipmentDetail(**base.model_dump(), items=items, orders=orders)


def _generate_remito_pdf(session: Session, remito: Remito, remito_items: list[RemitoItem], remito_type: str) -> str | None:
    if not remito_items:
        return None
    storage_dir = get_remitos_dir_new()
    storage_dir.mkdir(parents=True, exist_ok=True)
    filename = f"remito_{remito.id}_{remito_type.lower()}.pdf"
    file_path = storage_dir / filename

    pdf = canvas.Canvas(str(file_path), pagesize=A4)
    width, height = A4
    x_margin = 20 * mm
    y = height - 25 * mm

    type_label = "Productos terminados" if remito_type == "PT" else "No-PT (consumibles y otros)"
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(x_margin, y, f"Remito #{remito.id}")
    y -= 8 * mm
    pdf.setFont("Helvetica", 11)
    pdf.drawString(x_margin, y, f"Tipo: {type_label}")
    y -= 6 * mm
    pdf.drawString(x_margin, y, f"Fecha de emisión: {remito.issue_date.strftime('%d/%m/%Y')}")
    y -= 6 * mm
    pdf.drawString(x_margin, y, f"Destino: {remito.destination}")
    y -= 6 * mm
    source_deposit = session.get(Deposit, remito.source_deposit_id) if remito.source_deposit_id else None
    if source_deposit:
        pdf.drawString(x_margin, y, f"Origen: {source_deposit.name}")
        y -= 6 * mm

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(x_margin, y, "Detalle de ítems")
    y -= 8 * mm
    pdf.setFont("Helvetica", 10)

    header = ["Código", "Producto", "Cantidad"]
    col_widths = [35 * mm, 100 * mm, 25 * mm]
    pdf.drawString(x_margin, y, header[0])
    pdf.drawString(x_margin + col_widths[0], y, header[1])
    pdf.drawString(x_margin + col_widths[0] + col_widths[1], y, header[2])
    y -= 5 * mm
    pdf.line(x_margin, y, width - x_margin, y)
    y -= 6 * mm

    for item in remito_items:
        sku = session.get(SKU, item.sku_id)
        sku_code = sku.code if sku else str(item.sku_id)
        sku_name = sku.name if sku else f"SKU {item.sku_id}"
        if y < 25 * mm:
            pdf.showPage()
            y = height - 25 * mm
        pdf.drawString(x_margin, y, sku_code)
        pdf.drawString(x_margin + col_widths[0], y, sku_name)
        pdf.drawRightString(x_margin + col_widths[0] + col_widths[1] + col_widths[2], y, str(item.quantity))
        y -= 6 * mm

    pdf.save()
    return str(file_path)


def _format_pick_list_quantity(value: float | int | None) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _generate_pick_list_pdf(
    shipment: Shipment,
    items: list[ShipmentItem],
    orders: list[OrderSummaryRead],
    sku_map: dict[int, SKU],
    deposit_name: str | None,
    current_user: User | None,
) -> io.BytesIO:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    x_margin = 20 * mm
    y = height - 25 * mm

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(x_margin, y, f"Pick List – Envío #{shipment.id}")
    y -= 8 * mm
    pdf.setFont("Helvetica", 11)
    pdf.drawString(x_margin, y, f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC")
    y -= 6 * mm
    pdf.drawString(x_margin, y, f"Estado: {shipment.status}")
    y -= 6 * mm
    pdf.drawString(x_margin, y, f"Depósito destino: {deposit_name or shipment.deposit_id}")
    y -= 10 * mm

    if orders:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(x_margin, y, "Pedidos incluidos")
        y -= 6 * mm
        pdf.setFont("Helvetica", 10)
        for order in orders:
            if y < 25 * mm:
                pdf.showPage()
                y = height - 25 * mm
            pdf.drawString(x_margin, y, f"Pedido #{order.id}")
            y -= 5 * mm
        y -= 4 * mm

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(x_margin, y, "Ítems a preparar")
    y -= 8 * mm
    pdf.setFont("Helvetica", 10)

    headers = ["Código SKU", "Nombre", "Cantidad", "UoM", "OK", "Observaciones"]
    col_widths = [30 * mm, 70 * mm, 25 * mm, 15 * mm, 10 * mm, 30 * mm]
    pdf.drawString(x_margin, y, headers[0])
    pdf.drawString(x_margin + col_widths[0], y, headers[1])
    pdf.drawRightString(x_margin + col_widths[0] + col_widths[1] + col_widths[2], y, headers[2])
    pdf.drawString(x_margin + col_widths[0] + col_widths[1] + col_widths[2] + 4 * mm, y, headers[3])
    pdf.drawString(x_margin + col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3] + 8 * mm, y, headers[4])
    pdf.drawString(
        x_margin + col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3] + col_widths[4] + 6 * mm,
        y,
        headers[5],
    )
    y -= 5 * mm
    pdf.line(x_margin, y, width - x_margin, y)
    y -= 6 * mm

    for item in items:
        sku = sku_map.get(item.sku_id)
        sku_code = sku.code if sku else ""
        sku_name = sku.name if sku else f"SKU {item.sku_id}"
        quantity = _format_pick_list_quantity(item.quantity)
        unit = sku.unit.value if sku and isinstance(sku.unit, UnitOfMeasure) else (sku.unit if sku else "")
        if y < 25 * mm:
            pdf.showPage()
            y = height - 25 * mm
        pdf.drawString(x_margin, y, sku_code)
        pdf.drawString(x_margin + col_widths[0], y, sku_name[:40])
        pdf.drawRightString(x_margin + col_widths[0] + col_widths[1] + col_widths[2], y, quantity)
        pdf.drawString(x_margin + col_widths[0] + col_widths[1] + col_widths[2] + 4 * mm, y, unit)
        box_x = x_margin + col_widths[0] + col_widths[1] + col_widths[2] + col_widths[3] + 8 * mm
        pdf.rect(box_x, y - 2 * mm, 4 * mm, 4 * mm)
        y -= 6 * mm

    if current_user:
        footer_text = f"Generado por {current_user.full_name or current_user.email}"
        pdf.setFont("Helvetica", 9)
        pdf.drawString(x_margin, 15 * mm, footer_text)
    pdf.save()
    buffer.seek(0)
    return buffer


@public_router.post("/auth/login", tags=["auth"], response_model=TokenResponse)
def login(payload: LoginRequest, session: Session = Depends(get_session)) -> TokenResponse:
    user = session.exec(select(User).where(User.email == payload.username)).first()
    if not user or not verify_password(payload.password, user.hashed_password):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales inválidas")
    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuario inactivo")

    if is_legacy_hash(user.hashed_password) or needs_rehash(user.hashed_password):
        user.hashed_password = hash_password(payload.password)
        user.updated_at = datetime.utcnow()
        session.add(user)
        session.commit()
        session.refresh(user)

    token = create_access_token({"sub": str(user.id)})
    return TokenResponse(access_token=token, token_type="bearer", expires_in=settings.jwt_expires_minutes * 60)


@router.get("/auth/me", tags=["auth"], response_model=UserRead)
def auth_me(current_user: User = Depends(get_current_user), session: Session = Depends(get_session)) -> UserRead:
    return _map_user(current_user, session)


@public_router.get("/health", tags=["health"])
def health() -> dict[str, str]:
    return {"status": "ok", "version": "0.1.0"}


@router.get(
    "/roles",
    tags=["admin"],
    dependencies=[Depends(require_permissions("roles.view"))],
)
def list_roles(session: Session = Depends(get_session)) -> list[Role]:
    return session.exec(select(Role)).all()


@router.get(
    "/permissions",
    tags=["admin"],
    response_model=list[PermissionRead],
    dependencies=[Depends(require_permissions("roles.view"))],
)
def list_permissions(session: Session = Depends(get_session)) -> list[PermissionRead]:
    return session.exec(select(Permission)).all()


@router.get(
    "/roles/{role_id}/permissions",
    tags=["admin"],
    response_model=list[str],
    dependencies=[Depends(require_permissions("roles.view"))],
)
def list_role_permissions(role_id: int, session: Session = Depends(get_session)) -> list[str]:
    role = session.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rol no encontrado")
    permissions = session.exec(
        select(Permission.key)
        .join(RolePermission, RolePermission.permission_id == Permission.id)
        .where(RolePermission.role_id == role_id)
    ).all()
    return list(permissions)


@router.put(
    "/roles/{role_id}/permissions",
    tags=["admin"],
    response_model=list[str],
    dependencies=[Depends(require_permissions("roles.create_edit"))],
)
def update_role_permissions(
    role_id: int,
    payload: RolePermissionsUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> list[str]:
    role = session.get(Role, role_id)
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rol no encontrado")
    keys = {permission.strip().lower() for permission in payload.permissions}
    if not keys:
        keys = set()
    valid_permissions = session.exec(select(Permission).where(Permission.key.in_(keys))).all()
    if len(valid_permissions) != len(keys):
        missing = sorted(keys.difference({permission.key for permission in valid_permissions}))
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Permisos inexistentes: {', '.join(missing)}")
    existing = session.exec(select(RolePermission).where(RolePermission.role_id == role_id)).all()
    existing_map = {item.permission_id: item for item in existing}
    before_permissions = session.exec(
        select(Permission.key).join(RolePermission, RolePermission.permission_id == Permission.id).where(RolePermission.role_id == role_id)
    ).all()
    desired_ids = {permission.id for permission in valid_permissions}
    for permission_id, item in existing_map.items():
        if permission_id not in desired_ids:
            session.delete(item)
    for permission in valid_permissions:
        if permission.id not in existing_map:
            session.add(RolePermission(role_id=role_id, permission_id=permission.id))
    after_permissions = sorted(permission.key for permission in valid_permissions)
    audit_context = _build_audit_context(request)
    diff = {"permissions": {"before": sorted(before_permissions), "after": after_permissions}}
    changes = _build_audit_changes(
        "update",
        "roles",
        role_id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "roles", role_id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    return after_permissions


@router.get(
    "/users",
    tags=["admin"],
    response_model=list[UserRead],
    dependencies=[Depends(require_permissions("users.view"))],
)
def list_users(session: Session = Depends(get_session)) -> list[UserRead]:
    users = session.exec(select(User)).all()
    return [_map_user(user, session) for user in users]


@router.post(
    "/users",
    tags=["admin"],
    status_code=status.HTTP_201_CREATED,
    response_model=UserRead,
    dependencies=[Depends(require_permissions("users.create"))],
)
def create_user(
    payload: UserCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> UserRead:
    existing = session.exec(select(User).where(User.email == payload.email)).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El email ya está registrado")

    if payload.role_id:
        role = session.get(Role, payload.role_id)
        if not role:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rol inexistente")

    user = User(
        email=payload.email,
        full_name=payload.full_name,
        hashed_password=hash_password(payload.password),
        role_id=payload.role_id,
        is_active=payload.is_active,
    )
    session.add(user)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "users",
        user.id,
        diff=_create_diff_from_fields(
            {
                "email": user.email,
                "full_name": user.full_name,
                "role_id": user.role_id,
                "is_active": user.is_active,
            },
            ["email", "full_name", "role_id", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "users", user.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(user)
    return _map_user(user, session)


@router.put(
    "/users/{user_id}",
    tags=["admin"],
    response_model=UserRead,
    dependencies=[Depends(require_permissions("users.edit"))],
)
def update_user(
    user_id: int,
    payload: UserUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> UserRead:
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    if payload.is_active is False and _is_admin_account(user, session):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede desactivar el usuario admin")

    if payload.email and payload.email != user.email:
        duplicate = session.exec(select(User).where(User.email == payload.email)).first()
        if duplicate:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El email ya está registrado")

    if payload.role_id:
        role = session.get(Role, payload.role_id)
        if not role:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Rol inexistente")

    update_data = payload.model_dump(exclude_unset=True)
    password = update_data.pop("password", None)
    diff = _build_diff_from_payload(user, update_data)
    for field, value in update_data.items():
        setattr(user, field, value)
    if password:
        user.hashed_password = hash_password(password)
    user.updated_at = datetime.utcnow()
    session.add(user)
    session.flush()

    audit_context = _build_audit_context(request)
    if password:
        diff["password"] = {"before": "***", "after": "***"}
    if diff:
        changes = _build_audit_changes(
            "update",
            "users",
            user.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "users", user.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)

    session.commit()
    session.refresh(user)
    return _map_user(user, session)


@router.delete(
    "/users/{user_id}",
    tags=["admin"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("users.deactivate"))],
)
def delete_user(
    user_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    if _is_admin_account(user, session):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede eliminar el usuario admin")
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "delete",
        "users",
        user.id,
        diff=_delete_diff_from_fields(
            {
                "email": user.email,
                "full_name": user.full_name,
                "role_id": user.role_id,
                "is_active": user.is_active,
            },
            ["email", "full_name", "role_id", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "users", user.id, AuditAction.DELETE, current_user.id, changes, audit_context=audit_context)
    session.delete(user)
    session.commit()


@router.get("/units", tags=["catalogs"], response_model=list[UnitRead], dependencies=[Depends(require_permissions("units.view"))])
def list_units() -> list[UnitRead]:
    """Catálogo controlado de unidades de medida."""

    unit_labels = {
        UnitOfMeasure.UNIT: "Unidad",
        UnitOfMeasure.KG: "Kilogramo",
        UnitOfMeasure.G: "Gramo",
        UnitOfMeasure.L: "Litro",
        UnitOfMeasure.ML: "Mililitro",
        UnitOfMeasure.PACK: "Pack",
        UnitOfMeasure.BOX: "Caja",
        UnitOfMeasure.M: "Metro",
        UnitOfMeasure.CM: "Centímetro",
    }
    return [{"code": code, "label": unit_labels.get(code, code)} for code in UnitOfMeasure]


@router.get(
    "/sku-types",
    tags=["catalogs"],
    response_model=list[SKUTypeRead],
    dependencies=[Depends(require_permissions("sku_types.view"))],
)
def list_sku_types(include_inactive: bool = False, session: Session = Depends(get_session)) -> list[SKUTypeRead]:
    statement = select(SKUType)
    if not include_inactive:
        statement = statement.where(SKUType.is_active.is_(True))
    types = session.exec(statement.order_by(SKUType.code)).all()
    return [SKUTypeRead.model_validate(item) for item in types]


@router.post(
    "/sku-types",
    tags=["catalogs"],
    status_code=status.HTTP_201_CREATED,
    response_model=SKUTypeRead,
    dependencies=[Depends(require_permissions("sku_types.create_edit"))],
)
def create_sku_type(
    payload: SKUTypeCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SKUTypeRead:
    code = payload.code.strip().upper()
    duplicate = session.exec(select(SKUType).where(SKUType.code == code)).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe un tipo de SKU con ese código")
    record = SKUType(code=code, label=payload.label, is_active=payload.is_active)
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "sku_types",
        record.id,
        diff=_create_diff_from_fields(
            {"code": record.code, "label": record.label, "is_active": record.is_active},
            ["code", "label", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "sku_types", record.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return SKUTypeRead.model_validate(record)


@router.put(
    "/sku-types/{sku_type_id}",
    tags=["catalogs"],
    response_model=SKUTypeRead,
    dependencies=[Depends(require_permissions("sku_types.create_edit"))],
)
def update_sku_type(
    sku_type_id: int,
    payload: SKUTypeUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SKUTypeRead:
    sku_type = session.get(SKUType, sku_type_id)
    if not sku_type:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de SKU no encontrado")
    update_data = payload.model_dump(exclude_unset=True)
    diff = _build_diff_from_payload(sku_type, update_data)
    for field, value in update_data.items():
        setattr(sku_type, field, value)
    sku_type.updated_at = datetime.utcnow()
    session.add(sku_type)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "sku_types",
            sku_type.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "sku_types", sku_type.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(sku_type)
    return SKUTypeRead.model_validate(sku_type)


@router.delete(
    "/sku-types/{sku_type_id}",
    tags=["catalogs"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("sku_types.delete"))],
)
def delete_sku_type(
    sku_type_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    sku_type = session.get(SKUType, sku_type_id)
    if not sku_type:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de SKU no encontrado")
    in_use = session.exec(select(SKU.id).where(SKU.sku_type_id == sku_type_id)).first()
    audit_context = _build_audit_context(request)
    if in_use:
        before_status = sku_type.is_active
        sku_type.is_active = False
        sku_type.updated_at = datetime.utcnow()
        session.add(sku_type)
        diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
        changes = _build_audit_changes(
            "status",
            "sku_types",
            sku_type.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "sku_types", sku_type.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    else:
        changes = _build_audit_changes(
            "delete",
            "sku_types",
            sku_type.id,
            diff=_delete_diff_from_fields(
                {"code": sku_type.code, "label": sku_type.label, "is_active": sku_type.is_active},
                ["code", "label", "is_active"],
            ),
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "sku_types", sku_type.id, AuditAction.DELETE, current_user.id, changes, audit_context=audit_context)
        session.delete(sku_type)
    session.commit()


@router.get(
    "/mermas/types",
    tags=["mermas"],
    response_model=list[MermaTypeRead],
    dependencies=[Depends(require_permissions("mermas.view"))],
)
def list_merma_types(
    stage: MermaStage | None = None,
    include_inactive: bool = False,
    session: Session = Depends(get_session),
) -> list[MermaTypeRead]:
    statement = select(MermaType)
    if stage:
        statement = statement.where(MermaType.stage == stage)
    if not include_inactive:
        statement = statement.where(MermaType.is_active.is_(True))
    types = session.exec(statement.order_by(MermaType.stage, MermaType.label)).all()
    return [MermaTypeRead.model_validate(type_) for type_ in types]


@router.post(
    "/mermas/types",
    tags=["mermas"],
    status_code=status.HTTP_201_CREATED,
    response_model=MermaTypeRead,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def create_merma_type(
    payload: MermaTypeCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MermaTypeRead:
    duplicate = session.exec(
        select(MermaType).where(MermaType.stage == payload.stage, MermaType.code == payload.code)
    ).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe un tipo con ese código en la etapa")
    record = MermaType(**payload.model_dump())
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "merma_types",
        record.id,
        diff=_create_diff_from_fields(
            {
                "stage": record.stage,
                "code": record.code,
                "label": record.label,
                "is_active": record.is_active,
            },
            ["stage", "code", "label", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "merma_types", record.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return MermaTypeRead.model_validate(record)


@router.put(
    "/mermas/types/{type_id}",
    tags=["mermas"],
    response_model=MermaTypeRead,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def update_merma_type(
    type_id: int,
    payload: MermaTypeUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MermaTypeRead:
    record = session.get(MermaType, type_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de merma no encontrado")

    new_stage = payload.stage or record.stage
    duplicate = session.exec(
        select(MermaType).where(MermaType.stage == new_stage, MermaType.code == record.code, MermaType.id != type_id)
    ).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe un tipo con ese código en la etapa")

    update_data = payload.model_dump(exclude_unset=True)
    diff = _build_diff_from_payload(record, update_data)
    for field, value in update_data.items():
        setattr(record, field, value)
    record.updated_at = datetime.utcnow()
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "merma_types",
            record.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "merma_types", record.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return MermaTypeRead.model_validate(record)


@router.delete(
    "/mermas/types/{type_id}",
    tags=["mermas"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def delete_merma_type(
    type_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    record = session.get(MermaType, type_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de merma no encontrado")
    before_status = record.is_active
    record.is_active = False
    record.updated_at = datetime.utcnow()
    session.add(record)
    audit_context = _build_audit_context(request)
    diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "merma_types",
        record.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "merma_types", record.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)


@router.get(
    "/mermas/causes",
    tags=["mermas"],
    response_model=list[MermaCauseRead],
    dependencies=[Depends(require_permissions("mermas.view"))],
)
def list_merma_causes(
    stage: MermaStage | None = None,
    include_inactive: bool = False,
    session: Session = Depends(get_session),
) -> list[MermaCauseRead]:
    statement = select(MermaCause)
    if stage:
        statement = statement.where(MermaCause.stage == stage)
    if not include_inactive:
        statement = statement.where(MermaCause.is_active.is_(True))
    causes = session.exec(statement.order_by(MermaCause.stage, MermaCause.label)).all()
    return [MermaCauseRead.model_validate(cause) for cause in causes]


@router.post(
    "/mermas/causes",
    tags=["mermas"],
    status_code=status.HTTP_201_CREATED,
    response_model=MermaCauseRead,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def create_merma_cause(
    payload: MermaCauseCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MermaCauseRead:
    duplicate = session.exec(
        select(MermaCause).where(MermaCause.stage == payload.stage, MermaCause.code == payload.code)
    ).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe una causa con ese código en la etapa")
    record = MermaCause(**payload.model_dump())
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "merma_causes",
        record.id,
        diff=_create_diff_from_fields(
            {
                "stage": record.stage,
                "code": record.code,
                "label": record.label,
                "is_active": record.is_active,
            },
            ["stage", "code", "label", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "merma_causes", record.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return MermaCauseRead.model_validate(record)


@router.put(
    "/mermas/causes/{cause_id}",
    tags=["mermas"],
    response_model=MermaCauseRead,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def update_merma_cause(
    cause_id: int,
    payload: MermaCauseUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MermaCauseRead:
    record = session.get(MermaCause, cause_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Causa de merma no encontrada")

    new_stage = payload.stage or record.stage
    duplicate = session.exec(
        select(MermaCause).where(MermaCause.stage == new_stage, MermaCause.code == record.code, MermaCause.id != cause_id)
    ).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe una causa con ese código en la etapa")

    update_data = payload.model_dump(exclude_unset=True)
    diff = _build_diff_from_payload(record, update_data)
    for field, value in update_data.items():
        setattr(record, field, value)
    record.updated_at = datetime.utcnow()
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "merma_causes",
            record.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "merma_causes", record.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return MermaCauseRead.model_validate(record)


@router.delete(
    "/mermas/causes/{cause_id}",
    tags=["mermas"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def delete_merma_cause(
    cause_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    record = session.get(MermaCause, cause_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Causa de merma no encontrada")
    before_status = record.is_active
    record.is_active = False
    record.updated_at = datetime.utcnow()
    session.add(record)
    audit_context = _build_audit_context(request)
    diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "merma_causes",
        record.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "merma_causes", record.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)


@router.get("/skus", tags=["sku"], response_model=list[SKURead], dependencies=[Depends(require_permissions("skus.view"))])
def list_skus(
    sku_type_ids: list[int] | None = Query(None),
    tags: list[str] | None = Query(None, description="Alias para sku_type_code"),
    include_inactive: bool = False,
    search: str | None = None,
    session: Session = Depends(get_session),
) -> list[SKURead]:
    """Listado simple para bootstrap de catálogo."""
    statement = select(SKU).join(SKUType)
    if sku_type_ids:
        statement = statement.where(SKU.sku_type_id.in_(sku_type_ids))
    if tags:
        normalized_tags = [tag.upper() for tag in tags]
        statement = statement.where(SKUType.code.in_(normalized_tags))
    if not include_inactive:
        statement = statement.where(SKU.is_active.is_(True), SKUType.is_active.is_(True))
    if search:
        like = f"%{search.lower()}%"
        statement = statement.where((SKU.name.ilike(like)) | (SKU.code.ilike(like)))
    statement = statement.order_by(SKU.name, SKU.code)
    skus = session.exec(statement).all()
    return [_map_sku(item, session) for item in skus]


@router.get("/skus/export.xlsx", tags=["sku"], dependencies=[Depends(require_permissions("skus.view"))])
def export_skus(
    sku_type_ids: list[int] | None = Query(None),
    tags: list[str] | None = Query(None, description="Alias para sku_type_code"),
    include_inactive: bool = True,
    search: str | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    statement = select(SKU).join(SKUType)
    if sku_type_ids:
        statement = statement.where(SKU.sku_type_id.in_(sku_type_ids))
    if tags:
        normalized_tags = [tag.upper() for tag in tags]
        statement = statement.where(SKUType.code.in_(normalized_tags))
    if not include_inactive:
        statement = statement.where(SKU.is_active.is_(True), SKUType.is_active.is_(True))
    if search:
        like = f"%{search.lower()}%"
        statement = statement.where((SKU.name.ilike(like)) | (SKU.code.ilike(like)))
    statement = statement.order_by(SKU.name, SKU.code)
    skus = session.exec(statement).all()
    sku_type_ids_set = {sku.sku_type_id for sku in skus}
    sku_types = session.exec(select(SKUType).where(SKUType.id.in_(sku_type_ids_set))).all() if sku_type_ids_set else []
    sku_type_map = {sku_type.id: sku_type for sku_type in sku_types}
    sku_ids = [sku.id for sku in skus if sku.id]
    rules = (
        session.exec(select(SemiConversionRule).where(SemiConversionRule.sku_id.in_(sku_ids))).all() if sku_ids else []
    )
    rule_map = {rule.sku_id: rule for rule in rules}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SKUs"
    sheet.append(
        [
            "id",
            "code",
            "name",
            "sku_type_id",
            "sku_type_code",
            "sku_type_label",
            "unit",
            "secondary_unit",
            "units_per_kg",
            "notes",
            "is_active",
            "alert_green_min",
            "alert_yellow_min",
            "created_at",
            "updated_at",
        ]
    )
    for sku in skus:
        sku_type = sku_type_map.get(sku.sku_type_id)
        type_code = sku_type.code if sku_type else ""
        type_label = sku_type.label if sku_type else ""
        secondary_unit = UnitOfMeasure.UNIT if type_code == SKU_SEMI_CODE else None
        sku_id = sku.id
        units_per_kg = rule_map.get(sku_id).units_per_kg if sku_id and sku_id in rule_map else None
        sheet.append(
            [
                sku.id,
                sku.code,
                sku.name,
                sku.sku_type_id,
                type_code,
                type_label,
                sku.unit.value if isinstance(sku.unit, UnitOfMeasure) else sku.unit,
                secondary_unit.value if isinstance(secondary_unit, UnitOfMeasure) else secondary_unit,
                float(units_per_kg) if units_per_kg is not None else None,
                sku.notes,
                sku.is_active,
                sku.alert_green_min,
                sku.alert_yellow_min,
                sku.created_at.isoformat() if sku.created_at else None,
                sku.updated_at.isoformat() if sku.updated_at else None,
            ]
        )
    filename = f"skus_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return _build_excel_response(workbook, filename)


@router.post(
    "/skus",
    tags=["sku"],
    status_code=status.HTTP_201_CREATED,
    response_model=SKURead,
    dependencies=[Depends(require_permissions("skus.create"))],
)
def create_sku(
    payload: SKUCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SKURead:
    existing = session.exec(select(SKU).where(SKU.code == payload.code)).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El código ya existe",
        )

    sku_type = _get_sku_type_or_404(session, payload.sku_type_id)
    if not sku_type.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El tipo de SKU está inactivo")

    units_per_kg = payload.units_per_kg or 1
    unit = payload.unit
    if sku_type.code == SKU_SEMI_CODE:
        unit = UnitOfMeasure.KG
    if sku_type.code != SKU_SEMI_CODE:
        units_per_kg = None

    _validate_sku_alert_thresholds(payload.alert_green_min, payload.alert_yellow_min)

    sku = SKU(
        code=payload.code,
        name=payload.name,
        sku_type_id=payload.sku_type_id,
        unit=unit,
        notes=payload.notes,
        is_active=payload.is_active,
        alert_green_min=payload.alert_green_min,
        alert_yellow_min=payload.alert_yellow_min,
    )
    session.add(sku)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "skus",
        sku.id,
        diff=_create_diff_from_fields(
            {
                "code": sku.code,
                "name": sku.name,
                "sku_type_id": sku.sku_type_id,
                "unit": sku.unit,
                "notes": sku.notes,
                "is_active": sku.is_active,
                "alert_green_min": sku.alert_green_min,
                "alert_yellow_min": sku.alert_yellow_min,
            },
            ["code", "name", "sku_type_id", "unit", "notes", "is_active", "alert_green_min", "alert_yellow_min"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "skus", sku.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(sku)
    if sku_type.code == SKU_SEMI_CODE:
        _upsert_semi_conversion_rule(session, sku.id, units_per_kg)
        session.commit()
    return _map_sku(sku, session)


@router.put(
    "/skus/{sku_id}",
    tags=["sku"],
    response_model=SKURead,
    dependencies=[Depends(require_permissions("skus.edit"))],
)
def update_sku(
    sku_id: int,
    payload: SKUUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SKURead:
    sku = session.get(SKU, sku_id)
    if not sku:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU no encontrado")

    update_data = payload.model_dump(exclude_unset=True)
    sku_type_id = update_data.get("sku_type_id", sku.sku_type_id)
    sku_type = _get_sku_type_or_404(session, sku_type_id)
    if not sku_type.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El tipo de SKU está inactivo")

    units_per_kg = update_data.pop("units_per_kg", None)
    if sku_type.code == SKU_SEMI_CODE:
        update_data["unit"] = UnitOfMeasure.KG
    if units_per_kg is None and sku_type.code == SKU_SEMI_CODE:
        units_per_kg = _get_semi_units_per_kg(session, sku_id)

    effective_alert_green_min = update_data["alert_green_min"] if "alert_green_min" in update_data else sku.alert_green_min
    effective_alert_yellow_min = update_data["alert_yellow_min"] if "alert_yellow_min" in update_data else sku.alert_yellow_min
    _validate_sku_alert_thresholds(effective_alert_green_min, effective_alert_yellow_min)

    diff = _build_diff_from_payload(sku, update_data)
    for field, value in update_data.items():
        setattr(sku, field, value)
    sku.updated_at = datetime.utcnow()
    session.add(sku)
    session.flush()

    if sku_type.code == SKU_SEMI_CODE:
        _upsert_semi_conversion_rule(session, sku.id, units_per_kg or 1)
    else:
        _delete_semi_conversion_rule(session, sku.id)
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "skus",
            sku.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "skus", sku.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(sku)
    return _map_sku(sku, session)


@router.patch(
    "/stock/alerts/thresholds/{sku_id}",
    tags=["reports"],
    response_model=StockAlertThresholdRead,
    dependencies=[Depends(require_permissions("skus.edit"))],
)
def update_stock_alert_thresholds(
    sku_id: int,
    payload: StockAlertThresholdUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StockAlertThresholdRead:
    sku = session.get(SKU, sku_id)
    if not sku:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU no encontrado")
    update_data = payload.model_dump(exclude_unset=True)
    note = update_data.pop("note", None)
    effective_green = update_data["alert_green_min"] if "alert_green_min" in update_data else sku.alert_green_min
    effective_yellow = update_data["alert_yellow_min"] if "alert_yellow_min" in update_data else sku.alert_yellow_min
    _validate_sku_alert_thresholds(effective_green, effective_yellow)
    diff = _build_diff_from_payload(sku, update_data)
    for field, value in update_data.items():
        setattr(sku, field, value)
    sku.updated_at = datetime.utcnow()
    session.add(sku)
    audit_context = _build_audit_context(request)
    metadata = audit_context.get("metadata", {})
    if note:
        metadata = {**metadata, "note": note}
    if diff or note:
        changes = _build_audit_changes(
            "alert-thresholds",
            "skus",
            sku.id,
            diff=diff,
            metadata=metadata,
        )
        _log_audit(session, "skus", sku.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(sku)
    return StockAlertThresholdRead(sku_id=sku.id, alert_green_min=sku.alert_green_min, alert_yellow_min=sku.alert_yellow_min)


@router.delete(
    "/skus/{sku_id}",
    tags=["sku"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("skus.deactivate"))],
)
def delete_sku(
    sku_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    sku = session.get(SKU, sku_id)
    if not sku:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU no encontrado")
    in_use_reasons = []
    if session.exec(select(RecipeItem.id).where(RecipeItem.component_id == sku_id)).first():
        in_use_reasons.append("componentes de recetas")
    if session.exec(select(Recipe.id).where(Recipe.product_id == sku_id)).first():
        in_use_reasons.append("recetas de producto")
    if session.exec(select(ProductionLot.id).where(ProductionLot.sku_id == sku_id)).first():
        in_use_reasons.append("lotes de producción")
    if session.exec(select(StockMovement.id).where(StockMovement.sku_id == sku_id)).first():
        in_use_reasons.append("movimientos de stock")
    if session.exec(select(StockLevel.id).where(StockLevel.sku_id == sku_id)).first():
        in_use_reasons.append("niveles de stock")
    if session.exec(select(OrderItem.id).where(OrderItem.sku_id == sku_id)).first():
        in_use_reasons.append("pedidos")
    if session.exec(select(RemitoItem.id).where(RemitoItem.sku_id == sku_id)).first():
        in_use_reasons.append("remitos")
    if session.exec(select(InventoryCountItem.id).where(InventoryCountItem.sku_id == sku_id)).first():
        in_use_reasons.append("inventarios")
    if session.exec(select(MermaEvent.id).where(MermaEvent.sku_id == sku_id)).first():
        in_use_reasons.append("mermas")
    if in_use_reasons:
        reason_text = ", ".join(in_use_reasons)
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"No se puede eliminar el producto porque está referenciado en: {reason_text}.",
        )
    before_status = sku.is_active
    sku.is_active = False
    sku.updated_at = datetime.utcnow()
    session.add(sku)
    audit_context = _build_audit_context(request)
    diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "skus",
        sku.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "skus", sku.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()


@router.patch(
    "/skus/{sku_id}/status",
    tags=["sku"],
    response_model=SKURead,
    dependencies=[Depends(require_permissions("skus.deactivate"))],
)
def update_sku_status(
    sku_id: int,
    payload: StatusUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SKURead:
    sku = session.get(SKU, sku_id)
    if not sku:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU no encontrado")
    before = {"is_active": sku.is_active}
    if sku.is_active == payload.is_active:
        return _map_sku(sku, session)
    sku.is_active = payload.is_active
    sku.updated_at = datetime.utcnow()
    session.add(sku)
    audit_context = _build_audit_context(request)
    diff = _diff_fields(before, {"is_active": payload.is_active}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "skus",
        sku.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "skus", sku.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(sku)
    return _map_sku(sku, session)


@router.get("/deposits", tags=["deposits"], response_model=list[DepositRead], dependencies=[Depends(require_permissions("deposits.view"))])
def list_deposits(include_inactive: bool = False, session: Session = Depends(get_session)) -> list[Deposit]:
    statement = select(Deposit)
    if not include_inactive:
        statement = statement.where(Deposit.is_active.is_(True))
    return session.exec(statement.order_by(Deposit.name)).all()


@router.post(
    "/deposits",
    tags=["deposits"],
    status_code=status.HTTP_201_CREATED,
    response_model=DepositRead,
    dependencies=[Depends(require_permissions("deposits.create"))],
)
def create_deposit(
    payload: DepositCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> Deposit:
    existing = session.exec(select(Deposit).where(Deposit.name == payload.name)).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El depósito ya existe",
        )

    deposit = Deposit(**payload.model_dump())
    session.add(deposit)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "deposits",
        deposit.id,
        diff=_create_diff_from_fields(
            {
                "name": deposit.name,
                "location": deposit.location,
                "controls_lot": deposit.controls_lot,
                "is_store": deposit.is_store,
                "is_active": deposit.is_active,
            },
            ["name", "location", "controls_lot", "is_store", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "deposits", deposit.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(deposit)
    return deposit


@router.put(
    "/deposits/{deposit_id}",
    tags=["deposits"],
    response_model=DepositRead,
    dependencies=[Depends(require_permissions("deposits.edit"))],
)
def update_deposit(
    deposit_id: int,
    payload: DepositUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> Deposit:
    deposit = session.get(Deposit, deposit_id)
    if not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Depósito no encontrado")

    update_data = payload.model_dump(exclude_unset=True)
    diff = _build_diff_from_payload(deposit, update_data)
    for field, value in update_data.items():
        setattr(deposit, field, value)
    session.add(deposit)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "deposits",
            deposit.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "deposits", deposit.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(deposit)
    return deposit


@router.delete(
    "/deposits/{deposit_id}",
    tags=["deposits"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("deposits.deactivate"))],
)
def delete_deposit(
    deposit_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    deposit = session.get(Deposit, deposit_id)
    if not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Depósito no encontrado")
    in_use_reasons = []
    if session.exec(select(StockMovement.id).where(StockMovement.deposit_id == deposit_id)).first():
        in_use_reasons.append("movimientos de stock")
    if session.exec(select(StockLevel.id).where(StockLevel.deposit_id == deposit_id)).first():
        in_use_reasons.append("niveles de stock")
    if session.exec(select(ProductionLot.id).where(ProductionLot.deposit_id == deposit_id)).first():
        in_use_reasons.append("lotes de producción")
    if session.exec(select(InventoryCount.id).where(InventoryCount.deposit_id == deposit_id)).first():
        in_use_reasons.append("inventarios")
    if session.exec(select(Order.id).where(Order.destination_deposit_id == deposit_id)).first():
        in_use_reasons.append("pedidos")
    if session.exec(select(Remito.id).where((Remito.source_deposit_id == deposit_id) | (Remito.destination_deposit_id == deposit_id))).first():
        in_use_reasons.append("remitos")
    if session.exec(select(Shipment.id).where(Shipment.deposit_id == deposit_id)).first():
        in_use_reasons.append("envíos")
    if session.exec(select(MermaEvent.id).where(MermaEvent.deposit_id == deposit_id)).first():
        in_use_reasons.append("mermas")
    if in_use_reasons:
        reason_text = ", ".join(in_use_reasons)
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"No se puede eliminar el depósito porque está referenciado en: {reason_text}.",
        )
    before_status = deposit.is_active
    deposit.is_active = False
    deposit.updated_at = datetime.utcnow()
    session.add(deposit)
    audit_context = _build_audit_context(request)
    diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "deposits",
        deposit.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "deposits", deposit.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()


@router.patch(
    "/deposits/{deposit_id}/status",
    tags=["deposits"],
    response_model=DepositRead,
    dependencies=[Depends(require_permissions("deposits.deactivate"))],
)
def update_deposit_status(
    deposit_id: int,
    payload: StatusUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> Deposit:
    deposit = session.get(Deposit, deposit_id)
    if not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Depósito no encontrado")
    before = {"is_active": deposit.is_active}
    if deposit.is_active == payload.is_active:
        return deposit
    deposit.is_active = payload.is_active
    deposit.updated_at = datetime.utcnow()
    session.add(deposit)
    audit_context = _build_audit_context(request)
    diff = _diff_fields(before, {"is_active": payload.is_active}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "deposits",
        deposit.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "deposits", deposit.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(deposit)
    return deposit


@router.get(
    "/production-lines",
    tags=["mermas"],
    response_model=list[ProductionLineRead],
    dependencies=[Depends(require_permissions("production_lines.view"))],
)
def list_production_lines(session: Session = Depends(get_session)) -> list[ProductionLineRead]:
    lines = session.exec(select(ProductionLine).order_by(ProductionLine.name)).all()
    return [ProductionLineRead(id=line.id, name=line.name, is_active=line.is_active) for line in lines]


@router.post(
    "/production-lines",
    tags=["mermas"],
    status_code=status.HTTP_201_CREATED,
    response_model=ProductionLineRead,
    dependencies=[Depends(require_permissions("production_lines.create_edit"))],
)
def create_production_line(
    payload: ProductionLineCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ProductionLineRead:
    existing = session.exec(select(ProductionLine).where(ProductionLine.name == payload.name)).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe una línea con ese nombre")
    line = ProductionLine(name=payload.name, is_active=payload.is_active)
    session.add(line)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "production_lines",
        line.id,
        diff=_create_diff_from_fields({"name": line.name, "is_active": line.is_active}, ["name", "is_active"]),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "production_lines", line.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(line)
    return ProductionLineRead(id=line.id, name=line.name, is_active=line.is_active)


@router.put(
    "/production-lines/{line_id}",
    tags=["mermas"],
    response_model=ProductionLineRead,
    dependencies=[Depends(require_permissions("production_lines.create_edit"))],
)
def update_production_line(
    line_id: int,
    payload: ProductionLineUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ProductionLineRead:
    line = session.get(ProductionLine, line_id)
    if not line:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Línea no encontrada")
    if payload.name and payload.name != line.name:
        duplicate = session.exec(select(ProductionLine).where(ProductionLine.name == payload.name)).first()
        if duplicate:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe una línea con ese nombre")
    update_data = payload.model_dump(exclude_unset=True)
    diff = _build_diff_from_payload(line, update_data)
    for field, value in update_data.items():
        setattr(line, field, value)
    line.updated_at = datetime.utcnow()
    session.add(line)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "production_lines",
            line.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "production_lines", line.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(line)
    return ProductionLineRead(id=line.id, name=line.name, is_active=line.is_active)



def _map_recipe(recipe: Recipe, session: Session) -> RecipeRead:
    session.refresh(recipe, attribute_names=["items"])
    items = []
    for item in recipe.items:
        component = session.get(SKU, item.component_id)
        component_code = component.code if component else ""
        component_name = component.name if component else f"SKU {item.component_id}"
        session.refresh(component, attribute_names=["sku_type"]) if component else None
        if component and component.sku_type and component.sku_type.code == SKU_SEMI_CODE:
            component_unit = UnitOfMeasure.UNIT
        else:
            component_unit = component.unit if component else UnitOfMeasure.UNIT
        items.append(
            {
                "component_id": item.component_id,
                "quantity": item.quantity,
                "component_code": component_code,
                "component_name": component_name,
                "component_unit": component_unit,
            }
        )

    return RecipeRead(
        id=recipe.id,
        product_id=recipe.product_id,
        name=recipe.name,
        items=items,
        is_active=recipe.is_active,
    )


def _serialize_recipe_items_for_audit(items: list[RecipeItem]) -> list[dict]:
    return [{"component_id": item.component_id, "quantity": item.quantity} for item in items]


@router.get(
    "/recipes",
    tags=["recipes"],
    response_model=list[RecipeRead],
    dependencies=[Depends(require_permissions("recipes.view"))],
)
def list_recipes(include_inactive: bool = False, session: Session = Depends(get_session)) -> list[RecipeRead]:
    statement = select(Recipe)
    if not include_inactive:
        statement = statement.where(Recipe.is_active.is_(True))
    recipes = session.exec(statement).all()
    return [_map_recipe(recipe, session) for recipe in recipes]


@router.post(
    "/recipes",
    tags=["recipes"],
    status_code=status.HTTP_201_CREATED,
    response_model=RecipeRead,
    dependencies=[Depends(require_permissions("recipes.create"))],
)
def create_recipe(
    payload: RecipeCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> RecipeRead:
    product = session.get(SKU, payload.product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Producto no encontrado")

    session.refresh(product, attribute_names=["sku_type"])
    product_type = product.sku_type.code if product.sku_type else ""
    if product_type not in SKU_PRODUCTION_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Solo se pueden crear recetas para productos PT o SEMI",
        )

    if not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La receta debe tener al menos un componente")

    # Validate components exist
    component_ids = {item.component_id for item in payload.items}
    existing_components = session.exec(select(SKU.id).where(SKU.id.in_(component_ids))).all()
    if len(existing_components) != len(component_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún componente no existe")

    recipe = Recipe(product_id=payload.product_id, name=payload.name, is_active=payload.is_active)
    session.add(recipe)
    session.flush()  # assign id for FK

    for item in payload.items:
        session.add(
            RecipeItem(
                recipe_id=recipe.id,
                component_id=item.component_id,
                quantity=item.quantity,
            )
        )

    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "recipes",
        recipe.id,
        diff=_create_diff_from_fields(
            {
                "product_id": recipe.product_id,
                "name": recipe.name,
                "is_active": recipe.is_active,
                "items": [item.model_dump() for item in payload.items],
            },
            ["product_id", "name", "is_active", "items"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "recipes", recipe.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(recipe)
    return _map_recipe(recipe, session)


@router.put(
    "/recipes/{recipe_id}",
    tags=["recipes"],
    response_model=RecipeRead,
    dependencies=[Depends(require_permissions("recipes.edit"))],
)
def update_recipe(
    recipe_id: int,
    payload: RecipeUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> RecipeRead:
    recipe = session.get(Recipe, recipe_id)
    if not recipe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receta no encontrada")

    session.refresh(recipe, attribute_names=["items"])
    before_items = _serialize_recipe_items_for_audit(recipe.items)
    product = session.get(SKU, payload.product_id)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Producto no encontrado")
    session.refresh(product, attribute_names=["sku_type"])
    product_type = product.sku_type.code if product.sku_type else ""
    if product_type not in SKU_PRODUCTION_TYPES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se permiten PT o SEMI")

    if not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La receta debe tener al menos un componente")

    component_ids = {item.component_id for item in payload.items}
    existing_components = session.exec(select(SKU.id).where(SKU.id.in_(component_ids))).all()
    if len(existing_components) != len(component_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún componente no existe")

    update_fields = payload.model_dump(exclude_unset=True)
    items_payload = update_fields.pop("items", None)
    diff = _build_diff_from_payload(recipe, update_fields)
    if items_payload is not None and before_items != items_payload:
        diff["items"] = {"before": before_items, "after": items_payload}

    recipe.product_id = payload.product_id
    recipe.name = payload.name
    if payload.is_active is not None:
        recipe.is_active = payload.is_active

    # Replace items
    current_items = session.exec(select(RecipeItem).where(RecipeItem.recipe_id == recipe.id)).all()
    for item in current_items:
        session.delete(item)
    session.flush()

    for item in payload.items:
        session.add(RecipeItem(recipe_id=recipe.id, component_id=item.component_id, quantity=item.quantity))

    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "recipes",
            recipe.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "recipes", recipe.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(recipe)
    return _map_recipe(recipe, session)


@router.delete(
    "/recipes/{recipe_id}",
    tags=["recipes"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("recipes.deactivate"))],
)
def delete_recipe(
    recipe_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    recipe = session.get(Recipe, recipe_id)
    if not recipe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receta no encontrada")
    if session.exec(select(ProductionLot.id).where(ProductionLot.sku_id == recipe.product_id)).first():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="No se puede eliminar la receta porque ya fue usada en producción.",
        )
    before_status = recipe.is_active
    recipe.is_active = False
    recipe.updated_at = datetime.utcnow()
    session.add(recipe)
    audit_context = _build_audit_context(request)
    diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "recipes",
        recipe.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "recipes", recipe.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()


@router.patch(
    "/recipes/{recipe_id}/status",
    tags=["recipes"],
    response_model=RecipeRead,
    dependencies=[Depends(require_permissions("recipes.deactivate"))],
)
def update_recipe_status(
    recipe_id: int,
    payload: StatusUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> RecipeRead:
    recipe = session.get(Recipe, recipe_id)
    if not recipe:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Receta no encontrada")
    before = {"is_active": recipe.is_active}
    if recipe.is_active == payload.is_active:
        return _map_recipe(recipe, session)
    recipe.is_active = payload.is_active
    recipe.updated_at = datetime.utcnow()
    session.add(recipe)
    audit_context = _build_audit_context(request)
    diff = _diff_fields(before, {"is_active": payload.is_active}, ["is_active"])
    changes = _build_audit_changes(
        "status",
        "recipes",
        recipe.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "recipes", recipe.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(recipe)
    return _map_recipe(recipe, session)


@router.get(
    "/orders",
    tags=["orders"],
    response_model=list[OrderRead],
    dependencies=[Depends(require_permissions("orders.view"))],
)
def list_orders(
    status_filter: OrderStatus | None = None,
    destination_deposit_id: int | None = None,
    session: Session = Depends(get_session),
) -> list[OrderRead]:
    statement = select(Order)
    if status_filter:
        statement = statement.where(Order.status == status_filter)
    if destination_deposit_id:
        statement = statement.where(Order.destination_deposit_id == destination_deposit_id)
    orders = session.exec(statement.order_by(Order.created_at.desc())).all()
    return [_map_order(order, session) for order in orders]


@router.get(
    "/orders/{order_id}",
    tags=["orders"],
    response_model=OrderRead,
    dependencies=[Depends(require_permissions("orders.view"))],
)
def get_order(order_id: int, session: Session = Depends(get_session)) -> OrderRead:
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pedido no encontrado")
    return _map_order(order, session)


@router.post(
    "/orders",
    tags=["orders"],
    status_code=status.HTTP_201_CREATED,
    response_model=OrderRead,
    dependencies=[Depends(require_permissions("orders.create"))],
)
def create_order(
    payload: OrderCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrderRead:
    if not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El pedido debe tener al menos un ítem")
    if not payload.requested_by or not payload.requested_by.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Indica quién ingresa el pedido")
    destination = _ensure_store_destination(session, payload.destination_deposit_id)
    if payload.status not in {OrderStatus.DRAFT, OrderStatus.SUBMITTED}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Estado de pedido inválido")
    _validate_required_delivery_date(payload.required_delivery_date)

    items_payload = [item.model_dump() for item in payload.items]
    _validate_order_items(session, items_payload)

    order = Order(
        destination=destination.name,
        destination_deposit_id=destination.id,
        requested_for=payload.requested_for,
        required_delivery_date=payload.required_delivery_date,
        requested_by=payload.requested_by.strip() if payload.requested_by else None,
        status=payload.status,
        notes=payload.notes,
        plant_internal_note=payload.plant_internal_note,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
    )
    session.add(order)
    session.flush()

    for item in payload.items:
        session.add(
            OrderItem(
                order_id=order.id,
                sku_id=item.sku_id,
                quantity=int(item.quantity),
                current_stock=int(item.current_stock) if item.current_stock is not None else None,
            )
        )

    audit_context = _build_audit_context(request)
    audit_payload = {
        "destination_deposit_id": order.destination_deposit_id,
        "requested_for": order.requested_for,
        "required_delivery_date": order.required_delivery_date,
        "requested_by": order.requested_by,
        "status": order.status,
        "notes": order.notes,
        "plant_internal_note": order.plant_internal_note,
        "items": [_serialize_order_item_payload(item) for item in items_payload],
    }
    changes = _build_audit_changes(
        "create",
        "orders",
        order.id,
        diff=_create_diff_from_fields(
            audit_payload,
            [
                "destination_deposit_id",
                "requested_for",
                "required_delivery_date",
                "requested_by",
                "status",
                "notes",
                "plant_internal_note",
                "items",
            ],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "orders", order.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(order)
    return _map_order(order, session)


@router.put(
    "/orders/{order_id}",
    tags=["orders"],
    response_model=OrderRead,
    dependencies=[Depends(require_permissions("orders.edit"))],
)
def update_order(
    order_id: int,
    payload: OrderUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrderRead:
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pedido no encontrado")
    if _is_local_account(current_user, session) and order.status != OrderStatus.DRAFT:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="El pedido ya fue enviado. El local no puede modificarlo.",
        )
    is_admin = _is_admin_account(current_user, session)
    allowed_statuses = {OrderStatus.DRAFT} if not is_admin else {
        OrderStatus.DRAFT,
        OrderStatus.SUBMITTED,
        OrderStatus.PARTIALLY_PREPARED,
    }
    if order.status not in allowed_statuses:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede editar el pedido en este estado")
    if payload.items is not None and not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El pedido debe tener al menos un ítem")
    if payload.requested_by is not None and not payload.requested_by.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Indica quién ingresa el pedido")
    if payload.required_delivery_date is not None:
        _validate_required_delivery_date(payload.required_delivery_date)

    existing_items = session.exec(select(OrderItem).where(OrderItem.order_id == order.id)).all()
    before_items = [_serialize_order_item(item) for item in existing_items]

    destination = None
    if payload.destination_deposit_id is not None:
        destination = _ensure_store_destination(session, payload.destination_deposit_id)

    if payload.items is not None:
        items_payload = [item.model_dump() for item in payload.items]
        _validate_order_items(session, items_payload)

    update_data = payload.model_dump(exclude_unset=True)
    items = update_data.pop("items", None)
    destination_deposit_id = update_data.pop("destination_deposit_id", None)
    if "requested_by" in update_data and isinstance(update_data["requested_by"], str):
        update_data["requested_by"] = update_data["requested_by"].strip()
    if update_data.get("status"):
        _ensure_order_transition(order, update_data["status"])
    if update_data.get("status") == OrderStatus.CANCELLED:
        _ensure_order_cancel_allowed(session, order)
        if order.status == OrderStatus.SUBMITTED:
            _remove_order_from_draft_shipments(session, order.id)
    audit_fields = update_data.copy()
    if destination_deposit_id is not None:
        audit_fields["destination_deposit_id"] = destination_deposit_id
    diff = _build_diff_from_payload(order, audit_fields)
    if items is not None:
        after_items = [_serialize_order_item_payload(item) for item in items]
        if before_items != after_items:
            diff["items"] = {"before": before_items, "after": after_items}

    for field, value in update_data.items():
        setattr(order, field, value)
    if update_data.get("status") == OrderStatus.CANCELLED:
        order.cancelled_at = datetime.utcnow()
        order.cancelled_by_user_id = current_user.id
        order.cancelled_by_name = current_user.full_name
    if destination:
        order.destination = destination.name
        order.destination_deposit_id = destination.id
    order.updated_at = datetime.utcnow()
    order.updated_by_user_id = current_user.id
    session.add(order)
    session.flush()

    if items is not None:
        payload_by_sku = {item["sku_id"]: item for item in items}
        if len(payload_by_sku) != len(items):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Hay ítems duplicados en el pedido")

        order_item_ids = [item.id for item in existing_items]
        assigned_quantities = _get_assigned_quantities(session, order_item_ids)
        dispatched_quantities = _get_dispatched_quantities(session, order_item_ids)
        existing_by_sku = {item.sku_id: item for item in existing_items}
        for existing_item in existing_items:
            assigned = assigned_quantities.get(existing_item.id, 0.0)
            dispatched = dispatched_quantities.get(existing_item.id, 0.0)
            payload_item = payload_by_sku.get(existing_item.sku_id)
            if assigned > 0 and not payload_item:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No se pueden eliminar ítems ya asignados a un envío",
                )
            if not payload_item:
                continue
            payload_quantity = int(payload_item["quantity"])
            payload_stock = payload_item.get("current_stock")
            payload_stock_value = int(payload_stock) if payload_stock is not None else None
            if dispatched > 0:
                if payload_quantity != int(existing_item.quantity) or payload_stock_value != existing_item.current_stock:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="No se pueden modificar ítems ya despachados",
                    )
            elif assigned > 0 and payload_quantity < assigned:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="La cantidad no puede ser menor a la asignada a envíos",
                )

        for existing_item in existing_items:
            assigned = assigned_quantities.get(existing_item.id, 0.0)
            if assigned > 0:
                continue
            if existing_item.sku_id not in payload_by_sku:
                session.delete(existing_item)
        session.flush()

        for item in items:
            existing_item = existing_by_sku.get(item["sku_id"])
            if existing_item:
                assigned = assigned_quantities.get(existing_item.id, 0.0)
                dispatched = dispatched_quantities.get(existing_item.id, 0.0)
                if dispatched <= 0:
                    existing_item.quantity = int(item["quantity"])
                    existing_item.current_stock = int(item["current_stock"]) if item.get("current_stock") is not None else None
                    session.add(existing_item)
                continue
            session.add(
                OrderItem(
                    order_id=order.id,
                    sku_id=item["sku_id"],
                    quantity=int(item["quantity"]),
                    current_stock=int(item["current_stock"]) if item.get("current_stock") is not None else None,
                )
            )

    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "orders",
            order.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "orders", order.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(order)
    return _map_order(order, session)


@router.post(
    "/orders/{order_id}/status",
    tags=["orders"],
    response_model=OrderRead,
    dependencies=[Depends(require_permissions("orders.edit"))],
)
def update_order_status(
    order_id: int,
    payload: OrderStatusUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> OrderRead:
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pedido no encontrado")
    if payload.status == OrderStatus.SUBMITTED:
        items = session.exec(select(OrderItem).where(OrderItem.order_id == order.id)).all()
        if not items:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El pedido debe tener al menos un ítem")
        for item in items:
            if not _is_integer_value(item.quantity):
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="El pedido tiene cantidades con decimales. Corrige antes de enviar.",
                )
            if item.current_stock is None:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Stock actual requerido")
            if not _is_integer_value(item.current_stock):
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="El stock actual debe ser un número entero",
                )
    _ensure_order_transition(order, payload.status)
    if payload.status == OrderStatus.CANCELLED:
        _ensure_order_cancel_allowed(session, order)
        if order.status == OrderStatus.SUBMITTED:
            _remove_order_from_draft_shipments(session, order.id)
    before_status = order.status
    order.status = payload.status
    if payload.status == OrderStatus.CANCELLED:
        order.cancelled_at = datetime.utcnow()
        order.cancelled_by_user_id = current_user.id
        order.cancelled_by_name = current_user.full_name
    order.updated_at = datetime.utcnow()
    order.updated_by_user_id = current_user.id
    session.add(order)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "cancel" if payload.status == OrderStatus.CANCELLED else "status",
        "orders",
        order.id,
        diff=_status_diff("status", before_status, payload.status),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "orders", order.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(order)
    return _map_order(order, session)


@router.delete(
    "/orders/{order_id}",
    tags=["orders"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("orders.cancel"))],
)
def delete_order(
    order_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pedido no encontrado")
    if _is_local_account(current_user, session) and order.status != OrderStatus.DRAFT:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="El pedido ya fue enviado. El local no puede modificarlo.",
        )
    _ensure_order_transition(order, OrderStatus.CANCELLED)
    _ensure_order_cancel_allowed(session, order)
    if order.status == OrderStatus.SUBMITTED:
        _remove_order_from_draft_shipments(session, order.id)
    items = session.exec(select(OrderItem).where(OrderItem.order_id == order.id)).all()
    for item in items:
        session.delete(item)
    session.delete(order)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "delete",
        "orders",
        order_id,
        diff=_delete_diff_from_fields(
            {
                "destination_deposit_id": order.destination_deposit_id,
                "status": order.status,
                "requested_by": order.requested_by,
            },
            ["destination_deposit_id", "status", "requested_by"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "orders", order_id, AuditAction.DELETE, current_user.id, changes, audit_context=audit_context)
    session.commit()


@router.post(
    "/shipments",
    tags=["shipments"],
    status_code=status.HTTP_201_CREATED,
    response_model=ShipmentRead,
    dependencies=[Depends(require_permissions("remitos.create"))],
)
def create_shipment(
    payload: ShipmentCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentRead:
    destination = _ensure_store_destination(session, payload.deposit_id)
    shipment = Shipment(
        deposit_id=destination.id,
        estimated_delivery_date=payload.estimated_delivery_date,
        status=ShipmentStatus.DRAFT,
    )
    session.add(shipment)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "shipments",
        shipment.id,
        diff=_create_diff_from_fields(
            {
                "deposit_id": shipment.deposit_id,
                "estimated_delivery_date": shipment.estimated_delivery_date,
                "status": shipment.status,
            },
            ["deposit_id", "estimated_delivery_date", "status"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "shipments", shipment.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session)


@router.get(
    "/shipments",
    tags=["shipments"],
    response_model=list[ShipmentRead],
    dependencies=[Depends(require_permissions("remitos.view"))],
)
def list_shipments(
    deposit_id: int | None = None,
    status_filter: ShipmentStatus | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    session: Session = Depends(get_session),
) -> list[ShipmentRead]:
    statement = select(Shipment)
    if deposit_id:
        statement = statement.where(Shipment.deposit_id == deposit_id)
    if status_filter:
        statement = statement.where(Shipment.status == status_filter)
    if date_from:
        statement = statement.where(Shipment.estimated_delivery_date >= date_from)
    if date_to:
        statement = statement.where(Shipment.estimated_delivery_date <= date_to)
    shipments = session.exec(statement.order_by(Shipment.created_at.desc())).all()
    return [_map_shipment(shipment, session) for shipment in shipments]


@router.get(
    "/shipments/{shipment_id}",
    tags=["shipments"],
    response_model=ShipmentDetail,
    dependencies=[Depends(require_permissions("remitos.view"))],
)
def get_shipment(shipment_id: int, session: Session = Depends(get_session)) -> ShipmentDetail:
    shipment = _get_shipment_or_404(session, shipment_id)
    return _map_shipment(shipment, session, include_items=True)


@router.get(
    "/shipments/{shipment_id}/pick-list.pdf",
    tags=["shipments"],
    dependencies=[Depends(require_permissions("remitos.view"))],
)
def get_shipment_pick_list(
    shipment_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StreamingResponse:
    shipment = _get_shipment_or_404(session, shipment_id)
    session.refresh(shipment, attribute_names=["items", "deposit"])
    orders = _get_shipment_orders(session, shipment.id)
    items = session.exec(select(ShipmentItem).where(ShipmentItem.shipment_id == shipment.id)).all()
    order_item_ids = [item.order_item_id for item in items]
    order_items = session.exec(select(OrderItem).where(OrderItem.id.in_(order_item_ids))).all() if order_item_ids else []
    sku_ids = {order_item.sku_id for order_item in order_items}
    skus = session.exec(select(SKU).where(SKU.id.in_(sku_ids))).all() if sku_ids else []
    sku_map = {sku.id: sku for sku in skus}
    buffer = _generate_pick_list_pdf(shipment, items, orders, sku_map, shipment.deposit.name if shipment.deposit else None, current_user)
    filename = f"pick_list_shipment_{shipment.id}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/pdf", headers=headers)


@router.put(
    "/shipments/{shipment_id}",
    tags=["shipments"],
    response_model=ShipmentRead,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def update_shipment(
    shipment_id: int,
    payload: ShipmentUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentRead:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden editar envíos en borrador")

    updates: dict[str, object] = {}
    before: dict[str, object] = {}

    if payload.deposit_id is not None:
        destination = _ensure_store_destination(session, payload.deposit_id)
        if shipment.deposit_id != destination.id:
            items = session.exec(select(ShipmentItem).where(ShipmentItem.shipment_id == shipment.id)).all()
            if items:
                order_ids = {item.order_id for item in items}
                orders = session.exec(select(Order).where(Order.id.in_(order_ids))).all()
                invalid_orders = [order.id for order in orders if order.destination_deposit_id != destination.id]
                if invalid_orders:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="El envío ya tiene pedidos asociados a otro local",
                    )
            before["deposit_id"] = shipment.deposit_id
            updates["deposit_id"] = destination.id
            shipment.deposit_id = destination.id

    if payload.estimated_delivery_date is not None and payload.estimated_delivery_date != shipment.estimated_delivery_date:
        before["estimated_delivery_date"] = shipment.estimated_delivery_date
        updates["estimated_delivery_date"] = payload.estimated_delivery_date
        shipment.estimated_delivery_date = payload.estimated_delivery_date

    if not updates:
        return _map_shipment(shipment, session)

    shipment.updated_at = datetime.utcnow()
    shipment.updated_by_user_id = current_user.id
    session.add(shipment)
    audit_context = _build_audit_context(request)
    diff = {field: {"before": before.get(field), "after": updates.get(field)} for field in updates}
    if diff:
        changes = _build_audit_changes(
            "update",
            "shipments",
            shipment.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "shipments", shipment.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session)


@router.post(
    "/shipments/{shipment_id}/add-orders",
    tags=["shipments"],
    response_model=ShipmentDetail,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def add_orders_to_shipment(
    shipment_id: int,
    payload: ShipmentAddOrders,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentDetail:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden editar envíos en borrador")
    if not payload.order_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecciona al menos un pedido")

    orders = session.exec(select(Order).where(Order.id.in_(payload.order_ids))).all()
    if len(orders) != len(set(payload.order_ids)):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Algún pedido no existe")

    existing_items = session.exec(select(ShipmentItem).where(ShipmentItem.shipment_id == shipment.id)).all()
    before_order_ids = sorted({item.order_id for item in existing_items})
    existing_order_item_ids = {item.order_item_id for item in existing_items}

    order_item_ids = []
    order_items_by_order: dict[int, list[OrderItem]] = {}
    for order in orders:
        if order.status == OrderStatus.DRAFT:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El pedido {order.id} está en borrador y no puede incluirse en envíos.",
            )
        if order.status not in {
            OrderStatus.SUBMITTED,
            OrderStatus.PARTIALLY_PREPARED,
            OrderStatus.PREPARED,
            OrderStatus.PARTIALLY_DISPATCHED,
        }:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El pedido {order.id} no está en estado Enviado, Preparado o Parcialmente despachado",
            )
        if order.destination_deposit_id != shipment.deposit_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Los pedidos deben ser del mismo local")
        items = session.exec(select(OrderItem).where(OrderItem.order_id == order.id)).all()
        if not items:
            continue
        order_items_by_order[order.id] = items
        order_item_ids.extend([item.id for item in items])

    assigned_quantities = _get_assigned_quantities(session, order_item_ids)
    added_items = []
    for order_id, items in order_items_by_order.items():
        for item in items:
            remaining = float(item.quantity) - assigned_quantities.get(item.id, 0.0)
            if remaining <= 0:
                continue
            if item.id in existing_order_item_ids:
                continue
            shipment_item = ShipmentItem(
                shipment_id=shipment.id,
                order_id=order_id,
                order_item_id=item.id,
                quantity=int(remaining),
            )
            session.add(shipment_item)
            added_items.append({"order_item_id": item.id, "quantity": int(remaining), "order_id": order_id})

    audit_context = _build_audit_context(request)
    after_order_ids = sorted(set(before_order_ids).union(payload.order_ids))
    diff = {
        "order_ids": {"before": before_order_ids, "after": after_order_ids},
        "items": {"before": [], "after": added_items},
    }
    changes = _build_audit_changes(
        "update",
        "shipments",
        shipment.id,
        diff=diff,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "shipments", shipment.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    _sync_order_statuses(session, set(payload.order_ids), current_user, shipment_id=shipment.id, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session, include_items=True)


@router.post(
    "/shipments/{shipment_id}/items",
    tags=["shipments"],
    response_model=ShipmentDetail,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def update_shipment_items(
    shipment_id: int,
    payload: list[ShipmentItemUpdate],
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentDetail:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden editar envíos en borrador")
    if not payload:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay ítems para actualizar")

    items = session.exec(select(ShipmentItem).where(ShipmentItem.shipment_id == shipment.id)).all()
    item_map = {item.order_item_id: item for item in items}
    order_item_ids = list(item_map.keys())
    assigned_quantities = _get_assigned_quantities(session, order_item_ids, exclude_shipment_id=shipment.id)

    before_quantities = {item.order_item_id: item.quantity for item in items}
    changes = []

    for update in payload:
        if update.order_item_id not in item_map:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ítem no pertenece al envío")
        if update.quantity < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La cantidad no puede ser negativa")
        order_item = session.get(OrderItem, update.order_item_id)
        if not order_item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ítem de pedido no encontrado")
        remaining = float(order_item.quantity) - assigned_quantities.get(order_item.id, 0.0)
        if update.quantity > remaining:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La cantidad supera lo pendiente por preparar",
            )
        shipment_item = item_map[update.order_item_id]
        if update.quantity == 0:
            session.delete(shipment_item)
            changes.append(
                {
                    "order_item_id": update.order_item_id,
                    "before": shipment_item.quantity,
                    "after": 0,
                }
            )
            continue
        if shipment_item.quantity != update.quantity:
            changes.append(
                {
                    "order_item_id": update.order_item_id,
                    "before": shipment_item.quantity,
                    "after": update.quantity,
                }
            )
        shipment_item.quantity = update.quantity
        shipment_item.updated_at = datetime.utcnow()
        session.add(shipment_item)

    audit_context = _build_audit_context(request)
    after_quantities = before_quantities.copy()
    for change in changes:
        after_quantities[change["order_item_id"]] = change["after"]
    diff = {"items": {"before": before_quantities, "after": after_quantities}}
    if changes:
        changes_payload = _build_audit_changes(
            "update",
            "shipments",
            shipment.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(
            session,
            "shipments",
            shipment.id,
            AuditAction.UPDATE,
            current_user.id,
            changes_payload,
            audit_context=audit_context,
        )
    order_ids = {item.order_id for item in items}
    _sync_order_statuses(session, order_ids, current_user, shipment_id=shipment.id, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session, include_items=True)


@router.post(
    "/shipments/{shipment_id}/prep-items",
    tags=["shipments"],
    response_model=ShipmentDetail,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def prep_shipment_items(
    shipment_id: int,
    payload: ShipmentPrepItemsRequest,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentDetail:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden preparar envíos en borrador")
    if not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay ítems para actualizar")

    item_ids = {item.shipment_item_id for item in payload.items}
    items = session.exec(
        select(ShipmentItem).where(ShipmentItem.shipment_id == shipment.id, ShipmentItem.id.in_(item_ids))
    ).all()
    item_map = {item.id: item for item in items}
    if len(item_map) != len(item_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún ítem no pertenece al envío")

    changes = []
    for update in payload.items:
        shipment_item = item_map[update.shipment_item_id]
        if shipment_item.is_ready == update.is_ready:
            continue
        changes.append(
            {
                "shipment_item_id": update.shipment_item_id,
                "before": shipment_item.is_ready,
                "after": update.is_ready,
            }
        )
        shipment_item.is_ready = update.is_ready
        shipment_item.updated_at = datetime.utcnow()
        session.add(shipment_item)

    audit_context = _build_audit_context(request)
    if changes:
        before_list = [
            {"shipment_item_id": change["shipment_item_id"], "is_ready": change["before"]} for change in changes
        ]
        after_list = [
            {"shipment_item_id": change["shipment_item_id"], "is_ready": change["after"]} for change in changes
        ]
        diff = {"items_ready": {"before": before_list, "after": after_list}}
        changes_payload = _build_audit_changes(
            "update",
            "shipments",
            shipment.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(
            session,
            "shipments",
            shipment.id,
            AuditAction.UPDATE,
            current_user.id,
            changes_payload,
            audit_context=audit_context,
        )

    order_ids = {item.order_id for item in items}
    _sync_order_statuses(session, order_ids, current_user, shipment_id=shipment.id, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session, include_items=True)


@router.post(
    "/shipments/{shipment_id}/cancel",
    tags=["shipments"],
    response_model=ShipmentRead,
    dependencies=[Depends(require_permissions("remitos.cancel"))],
)
def cancel_shipment(
    shipment_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentRead:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden cancelar envíos en borrador")

    shipment_payload = _map_shipment(shipment, session)
    items = session.exec(select(ShipmentItem).where(ShipmentItem.shipment_id == shipment.id)).all()
    order_ids = {item.order_id for item in items}
    for item in items:
        session.delete(item)

    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "cancel",
        "shipments",
        shipment.id,
        diff=_status_diff("status", shipment.status, "cancelled"),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "shipments", shipment.id, AuditAction.CANCEL, current_user.id, changes, audit_context=audit_context)
    session.delete(shipment)
    session.flush()
    _sync_order_statuses(session, order_ids, current_user, shipment_id=shipment.id, audit_context=audit_context)
    session.commit()
    return shipment_payload


@router.post(
    "/shipments/{shipment_id}/confirm",
    tags=["shipments"],
    response_model=ShipmentDetail,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def confirm_shipment(
    shipment_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentDetail:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El envío ya fue confirmado")
    session.refresh(shipment, attribute_names=["items"])
    if not shipment.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El envío no tiene ítems cargados")

    order_item_ids = [item.order_item_id for item in shipment.items]
    assigned_quantities = _get_assigned_quantities(session, order_item_ids, exclude_shipment_id=shipment.id)

    for item in shipment.items:
        order_item = session.get(OrderItem, item.order_item_id)
        if not order_item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ítem de pedido no encontrado")
        remaining = float(order_item.quantity) - assigned_quantities.get(order_item.id, 0.0)
        if item.quantity <= 0 or item.quantity > remaining:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La cantidad a despachar es inválida para algún ítem",
            )

    source_deposit = _default_source_deposit(session)
    movement_type = _get_movement_type_by_code(session, "REMITO")

    remitos_created = []
    stock_movement_ids = []
    remito_items_by_type: dict[str, list[ShipmentItem]] = {"PT": [], "NO_PT": []}
    for item in shipment.items:
        order_item = session.get(OrderItem, item.order_item_id)
        sku = session.get(SKU, order_item.sku_id) if order_item else None
        sku_type_code = sku.sku_type.code if sku and sku.sku_type else ""
        key = "PT" if sku_type_code == "PT" else "NO_PT"
        remito_items_by_type[key].append(item)

    audit_context = _build_audit_context(request)
    for remito_type, items in remito_items_by_type.items():
        if not items:
            continue
        destination_deposit = _get_deposit_or_404(session, shipment.deposit_id)
        remito = Remito(
            shipment_id=shipment.id,
            status=RemitoStatus.DISPATCHED,
            destination=destination_deposit.name,
            issue_date=date.today(),
            source_deposit_id=source_deposit.id,
            destination_deposit_id=destination_deposit.id,
            dispatched_at=datetime.utcnow(),
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
        )
        session.add(remito)
        session.flush()

        remito_items = []
        for shipment_item in items:
            order_item = session.get(OrderItem, shipment_item.order_item_id)
            if not order_item:
                continue
            remito_item = RemitoItem(
                remito_id=remito.id,
                sku_id=order_item.sku_id,
                quantity=shipment_item.quantity,
            )
            session.add(remito_item)
            remito_items.append(remito_item)
        session.flush()

        remito.pdf_path = _generate_remito_pdf(session, remito, remito_items, remito_type)
        remito.updated_at = datetime.utcnow()
        remitos_created.append(remito.id)
        remito_changes = _build_audit_changes(
            "create",
            "remitos",
            remito.id,
            diff=_create_diff_from_fields(
                {
                    "shipment_id": shipment.id,
                    "status": remito.status,
                    "source_deposit_id": remito.source_deposit_id,
                    "destination_deposit_id": remito.destination_deposit_id,
                    "issue_date": remito.issue_date,
                    "remito_type": remito_type,
                },
                [
                    "shipment_id",
                    "status",
                    "source_deposit_id",
                    "destination_deposit_id",
                    "issue_date",
                    "remito_type",
                ],
            ),
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "remitos", remito.id, AuditAction.CREATE, current_user.id, remito_changes, audit_context=audit_context)

    for item in shipment.items:
        order_item = session.get(OrderItem, item.order_item_id)
        if not order_item:
            continue
        movement_payload = StockMovementCreate(
            sku_id=order_item.sku_id,
            deposit_id=source_deposit.id,
            movement_type_id=movement_type.id,
            quantity=item.quantity,
            is_outgoing=True,
            reference_type="SHIPMENT",
            reference_id=shipment.id,
            reference_item_id=item.id,
            reference=f"ENVIO-{shipment.id}",
            created_by_user_id=current_user.id,
        )
        _, movement = _apply_stock_movement(
            session,
            movement_payload,
            allow_negative_balance=True,
            audit_context=audit_context,
            audit_movement=True,
        )
        stock_movement_ids.append(movement.id)

    shipment.status = ShipmentStatus.CONFIRMED
    shipment.updated_at = datetime.utcnow()
    session.add(shipment)
    session.flush()

    _sync_order_statuses(
        session,
        {item.order_id for item in shipment.items},
        current_user,
        shipment_id=shipment.id,
        audit_context=audit_context,
    )

    refs = [{"type": "remitos", "id": remito_id} for remito_id in remitos_created]
    refs.extend({"type": "stock_movements", "id": movement_id} for movement_id in stock_movement_ids)
    changes = _build_audit_changes(
        "status",
        "shipments",
        shipment.id,
        diff=_status_diff("status", ShipmentStatus.DRAFT, ShipmentStatus.CONFIRMED),
        refs=refs,
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "shipments", shipment.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session, include_items=True)


@router.post(
    "/shipments/{shipment_id}/dispatch",
    tags=["shipments"],
    response_model=ShipmentRead,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def dispatch_shipment(
    shipment_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> ShipmentRead:
    shipment = _get_shipment_or_404(session, shipment_id)
    if shipment.status != ShipmentStatus.CONFIRMED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden despachar envíos confirmados")
    shipment.status = ShipmentStatus.DISPATCHED
    shipment.updated_at = datetime.utcnow()
    session.add(shipment)
    session.flush()
    audit_context = _build_audit_context(request)
    _sync_order_statuses(
        session,
        {item.order_id for item in shipment.items},
        current_user,
        shipment_id=shipment.id,
        audit_context=audit_context,
    )
    changes = _build_audit_changes(
        "status",
        "shipments",
        shipment.id,
        diff=_status_diff("status", ShipmentStatus.CONFIRMED, ShipmentStatus.DISPATCHED),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "shipments", shipment.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(shipment)
    return _map_shipment(shipment, session)

@router.get(
    "/remitos",
    tags=["remitos"],
    response_model=list[RemitoRead],
    dependencies=[Depends(require_permissions("remitos.view"))],
)
def list_remitos(
    status_filter: RemitoStatus | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    session: Session = Depends(get_session),
) -> list[RemitoRead]:
    statement = select(Remito)
    if status_filter:
        statement = statement.where(Remito.status == status_filter)
    if date_from:
        statement = statement.where(Remito.issue_date >= date_from)
    if date_to:
        statement = statement.where(Remito.issue_date <= date_to)
    remitos = session.exec(statement.order_by(Remito.created_at.desc())).all()
    return [_map_remito(remito, session) for remito in remitos]


@router.get(
    "/remitos/{remito_id}",
    tags=["remitos"],
    response_model=RemitoRead,
    dependencies=[Depends(require_permissions("remitos.view"))],
)
def get_remito(remito_id: int, session: Session = Depends(get_session)) -> RemitoRead:
    remito = _get_remito_or_404(session, remito_id)
    return _map_remito(remito, session)


@router.get(
    "/remitos/{remito_id}/pdf",
    tags=["remitos"],
    response_class=FileResponse,
    dependencies=[Depends(require_permissions("remitos.view"))],
)
def get_remito_pdf(remito_id: int, session: Session = Depends(get_session)) -> FileResponse:
    remito = _get_remito_or_404(session, remito_id)
    if not remito.pdf_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El remito no tiene PDF generado")
    pdf_path = resolve_remito_pdf_path(remito.id, remito.pdf_path)
    if not pdf_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No se encontró el PDF del remito")
    return FileResponse(pdf_path, media_type="application/pdf", filename=pdf_path.name)


@router.post(
    "/remitos/from-order/{order_id}",
    tags=["remitos"],
    response_model=RemitoRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_permissions("remitos.create"))],
)
def create_remito_from_order(
    order_id: int,
    session: Session = Depends(get_session),
) -> RemitoRead:
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Los remitos se generan exclusivamente desde envíos confirmados",
    )


@router.post(
    "/remitos/{remito_id}/dispatch",
    tags=["remitos"],
    response_model=RemitoRead,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def dispatch_remito(
    remito_id: int,
    request: Request,
    payload: RemitoDispatchRequest | None = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> RemitoRead:
    remito = _get_remito_or_404(session, remito_id)
    session.refresh(remito, attribute_names=["items"])
    if remito.shipment_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Los remitos de envíos confirmados no se despachan manualmente",
        )
    if remito.status in {RemitoStatus.DISPATCHED, RemitoStatus.RECEIVED, RemitoStatus.CANCELLED}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El remito ya fue procesado")
    if not remito.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El remito no tiene ítems cargados")

    source_deposit = _get_deposit_or_404(session, remito.source_deposit_id) if remito.source_deposit_id else _default_source_deposit(session)
    movement_type = _get_movement_type_by_code(session, "REMITO")
    reference = f"REMITO-{remito.id}"
    movement_date = payload.movement_date if payload else None

    audit_context = _build_audit_context(request)
    for item in remito.items:
        sku = session.get(SKU, item.sku_id)
        if not sku:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"SKU {item.sku_id} no encontrado")
        if source_deposit.controls_lot:
            consumptions = _calculate_consumption_by_lot(
                session, item.sku_id, source_deposit.id, item.quantity, strict=True
            )
            for lot, quantity in consumptions:
                movement_payload = StockMovementCreate(
                    sku_id=item.sku_id,
                    deposit_id=source_deposit.id,
                    movement_type_id=movement_type.id,
                    quantity=quantity,
                    is_outgoing=True,
                    reference_type="REMITO",
                    reference_id=remito.id,
                    reference_item_id=item.id,
                    reference=reference,
                    lot_code=lot.lot_code if lot else None,
                    production_lot_id=lot.id if lot else None,
                    movement_date=movement_date,
                    created_by_user_id=current_user.id,
                )
                _apply_stock_movement(
                    session,
                    movement_payload,
                    allow_negative_balance=False,
                    audit_context=audit_context,
                    audit_movement=True,
                )
        else:
            movement_payload = StockMovementCreate(
                sku_id=item.sku_id,
                deposit_id=source_deposit.id,
                movement_type_id=movement_type.id,
                quantity=item.quantity,
                is_outgoing=True,
                reference_type="REMITO",
                reference_id=remito.id,
                reference_item_id=item.id,
                reference=reference,
                movement_date=movement_date,
                created_by_user_id=current_user.id,
            )
            _apply_stock_movement(
                session,
                movement_payload,
                allow_negative_balance=False,
                audit_context=audit_context,
                audit_movement=True,
            )

    before_status = remito.status
    remito.status = RemitoStatus.DISPATCHED
    remito.dispatched_at = datetime.utcnow()
    remito.source_deposit_id = source_deposit.id
    remito.updated_at = datetime.utcnow()
    remito.updated_by_user_id = current_user.id
    session.add(remito)
    changes = _build_audit_changes(
        "status",
        "remitos",
        remito.id,
        diff=_status_diff("status", before_status, RemitoStatus.DISPATCHED),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "remitos", remito.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(remito)
    return _map_remito(remito, session)


@router.post(
    "/remitos/{remito_id}/receive",
    tags=["remitos"],
    response_model=RemitoRead,
    dependencies=[Depends(require_permissions("remitos.edit"))],
)
def receive_remito(
    remito_id: int,
    request: Request,
    payload: RemitoReceiveRequest | None = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> RemitoRead:
    remito = _get_remito_or_404(session, remito_id)
    session.refresh(remito, attribute_names=["items"])
    if remito.shipment_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Los remitos de envíos confirmados no se reciben manualmente",
        )
    if remito.status is not RemitoStatus.DISPATCHED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden recibir remitos despachados")
    if not remito.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El remito no tiene ítems cargados")

    destination_deposit = (
        _get_deposit_or_404(session, remito.destination_deposit_id)
        if remito.destination_deposit_id
        else _default_source_deposit(session)
    )
    movement_type = _get_movement_type_by_code(session, "REMITO")
    reference = f"REMITO-{remito.id}"
    movement_date = payload.movement_date if payload else None

    audit_context = _build_audit_context(request)
    for item in remito.items:
        movement_payload = StockMovementCreate(
            sku_id=item.sku_id,
            deposit_id=destination_deposit.id,
            movement_type_id=movement_type.id,
            quantity=item.quantity,
            is_outgoing=False,
            reference_type="REMITO",
            reference_id=remito.id,
            reference_item_id=item.id,
            reference=f"{reference}-RECIBIDO",
            movement_date=movement_date,
            created_by_user_id=current_user.id,
        )
        _apply_stock_movement(
            session,
            movement_payload,
            allow_negative_balance=True,
            audit_context=audit_context,
            audit_movement=True,
        )

    before_status = remito.status
    remito.status = RemitoStatus.RECEIVED
    remito.received_at = datetime.utcnow()
    remito.destination_deposit_id = destination_deposit.id
    remito.updated_at = datetime.utcnow()
    remito.updated_by_user_id = current_user.id
    session.add(remito)
    changes = _build_audit_changes(
        "status",
        "remitos",
        remito.id,
        diff=_status_diff("status", before_status, RemitoStatus.RECEIVED),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "remitos", remito.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(remito)
    return _map_remito(remito, session)


@router.post(
    "/remitos/{remito_id}/cancel",
    tags=["remitos"],
    response_model=RemitoRead,
    dependencies=[Depends(require_permissions("remitos.cancel"))],
)
def cancel_remito(
    remito_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> RemitoRead:
    remito = _get_remito_or_404(session, remito_id)
    if remito.shipment_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Los remitos de envíos confirmados no se pueden cancelar manualmente",
        )
    if remito.status in {RemitoStatus.DISPATCHED, RemitoStatus.RECEIVED, RemitoStatus.CANCELLED}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede cancelar un remito procesado")
    before_status = remito.status
    remito.status = RemitoStatus.CANCELLED
    remito.cancelled_at = datetime.utcnow()
    remito.updated_at = datetime.utcnow()
    remito.updated_by_user_id = current_user.id
    session.add(remito)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "cancel",
        "remitos",
        remito.id,
        diff=_status_diff("status", before_status, RemitoStatus.CANCELLED),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "remitos", remito.id, AuditAction.CANCEL, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(remito)
    return _map_remito(remito, session)


@router.get(
    "/stock/movement-types",
    tags=["stock"],
    response_model=list[StockMovementTypeRead],
    dependencies=[Depends(require_permissions("movement_types.view"))],
)
def list_stock_movement_types(
    include_inactive: bool = False, session: Session = Depends(get_session)
) -> list[StockMovementTypeRead]:
    statement = select(StockMovementType)
    if not include_inactive:
        statement = statement.where(StockMovementType.is_active.is_(True))
    types = session.exec(statement.order_by(StockMovementType.code)).all()
    return [StockMovementTypeRead.model_validate(item) for item in types]


@router.post(
    "/stock/movement-types",
    tags=["stock"],
    status_code=status.HTTP_201_CREATED,
    response_model=StockMovementTypeRead,
    dependencies=[Depends(require_permissions("movement_types.create_edit"))],
)
def create_stock_movement_type(
    payload: StockMovementTypeCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StockMovementTypeRead:
    code = payload.code.strip().upper()
    duplicate = session.exec(select(StockMovementType).where(StockMovementType.code == code)).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe un tipo de movimiento con ese código")
    record = StockMovementType(code=code, label=payload.label, is_active=payload.is_active)
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "stock_movement_types",
        record.id,
        diff=_create_diff_from_fields(
            {"code": record.code, "label": record.label, "is_active": record.is_active},
            ["code", "label", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(
        session,
        "stock_movement_types",
        record.id,
        AuditAction.CREATE,
        current_user.id,
        changes,
        audit_context=audit_context,
    )
    session.commit()
    session.refresh(record)
    return StockMovementTypeRead.model_validate(record)


@router.put(
    "/stock/movement-types/{movement_type_id}",
    tags=["stock"],
    response_model=StockMovementTypeRead,
    dependencies=[Depends(require_permissions("movement_types.create_edit"))],
)
def update_stock_movement_type(
    movement_type_id: int,
    payload: StockMovementTypeUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StockMovementTypeRead:
    record = session.get(StockMovementType, movement_type_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de movimiento no encontrado")
    update_data = payload.model_dump(exclude_unset=True)
    diff = _build_diff_from_payload(record, update_data)
    for field, value in update_data.items():
        setattr(record, field, value)
    record.updated_at = datetime.utcnow()
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "stock_movement_types",
            record.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(
            session,
            "stock_movement_types",
            record.id,
            AuditAction.UPDATE,
            current_user.id,
            changes,
            audit_context=audit_context,
        )
    session.commit()
    session.refresh(record)
    return StockMovementTypeRead.model_validate(record)


@router.delete(
    "/stock/movement-types/{movement_type_id}",
    tags=["stock"],
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions("movement_types.delete"))],
)
def delete_stock_movement_type(
    movement_type_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> None:
    record = session.get(StockMovementType, movement_type_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de movimiento no encontrado")
    in_use = session.exec(select(StockMovement.id).where(StockMovement.movement_type_id == movement_type_id)).first()
    audit_context = _build_audit_context(request)
    if in_use:
        before_status = record.is_active
        record.is_active = False
        record.updated_at = datetime.utcnow()
        session.add(record)
        diff = _diff_fields({"is_active": before_status}, {"is_active": False}, ["is_active"])
        changes = _build_audit_changes(
            "status",
            "stock_movement_types",
            record.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(
            session,
            "stock_movement_types",
            record.id,
            AuditAction.STATUS,
            current_user.id,
            changes,
            audit_context=audit_context,
        )
    else:
        changes = _build_audit_changes(
            "delete",
            "stock_movement_types",
            record.id,
            diff=_delete_diff_from_fields(
                {"code": record.code, "label": record.label, "is_active": record.is_active},
                ["code", "label", "is_active"],
            ),
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(
            session,
            "stock_movement_types",
            record.id,
            AuditAction.DELETE,
            current_user.id,
            changes,
            audit_context=audit_context,
        )
        session.delete(record)
    session.commit()


@router.get(
    "/suppliers",
    tags=["purchases"],
    response_model=list[SupplierRead],
    dependencies=[Depends(require_permissions("suppliers.view"))],
)
def list_suppliers(include_inactive: bool = False, session: Session = Depends(get_session)) -> list[SupplierRead]:
    statement = select(Supplier)
    if not include_inactive:
        statement = statement.where(Supplier.is_active.is_(True))
    records = session.exec(statement.order_by(Supplier.name)).all()
    return [_map_supplier(record) for record in records]


@router.post(
    "/suppliers",
    tags=["purchases"],
    status_code=status.HTTP_201_CREATED,
    response_model=SupplierRead,
    dependencies=[Depends(require_permissions("suppliers.create"))],
)
def create_supplier(
    payload: SupplierCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SupplierRead:
    duplicate = session.exec(select(Supplier).where(Supplier.name == payload.name.strip())).first()
    if duplicate:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe un proveedor con ese nombre")
    record = Supplier(
        name=payload.name.strip(),
        tax_id=payload.tax_id,
        email=payload.email,
        phone=payload.phone,
        is_active=payload.is_active,
    )
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "suppliers",
        record.id,
        diff=_create_diff_from_fields(
            {
                "name": record.name,
                "tax_id": record.tax_id,
                "email": record.email,
                "phone": record.phone,
                "is_active": record.is_active,
            },
            ["name", "tax_id", "email", "phone", "is_active"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "suppliers", record.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return _map_supplier(record)


@router.put(
    "/suppliers/{supplier_id}",
    tags=["purchases"],
    response_model=SupplierRead,
    dependencies=[Depends(require_permissions("suppliers.edit"))],
)
def update_supplier(
    supplier_id: int,
    payload: SupplierUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> SupplierRead:
    record = session.get(Supplier, supplier_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proveedor no encontrado")
    update_data = payload.model_dump(exclude_unset=True)
    if "name" in update_data and update_data["name"]:
        update_data["name"] = update_data["name"].strip()
        duplicate = session.exec(
            select(Supplier).where(Supplier.name == update_data["name"], Supplier.id != supplier_id)
        ).first()
        if duplicate:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ya existe un proveedor con ese nombre")
    diff = _build_diff_from_payload(record, update_data)
    for field, value in update_data.items():
        setattr(record, field, value)
    record.updated_at = datetime.utcnow()
    session.add(record)
    session.flush()
    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "suppliers",
            record.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "suppliers", record.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(record)
    return _map_supplier(record)


@router.get(
    "/purchases/receipts",
    tags=["purchases"],
    response_model=list[PurchaseReceiptRead],
    dependencies=[Depends(require_permissions("purchases.view"))],
)
def list_purchase_receipts(
    supplier_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    session: Session = Depends(get_session),
) -> list[PurchaseReceiptRead]:
    statement = select(PurchaseReceipt)
    if supplier_id:
        statement = statement.where(PurchaseReceipt.supplier_id == supplier_id)
    if date_from:
        statement = statement.where(PurchaseReceipt.received_at >= date_from)
    if date_to:
        statement = statement.where(PurchaseReceipt.received_at <= date_to)
    receipts = session.exec(statement.order_by(PurchaseReceipt.received_at.desc(), PurchaseReceipt.id.desc())).all()
    return [_map_purchase_receipt(receipt, session) for receipt in receipts]


@router.get(
    "/purchases/receipts/{receipt_id}",
    tags=["purchases"],
    response_model=PurchaseReceiptRead,
    dependencies=[Depends(require_permissions("purchases.view"))],
)
def get_purchase_receipt(receipt_id: int, session: Session = Depends(get_session)) -> PurchaseReceiptRead:
    receipt = session.get(PurchaseReceipt, receipt_id)
    if not receipt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ingreso de compra no encontrado")
    return _map_purchase_receipt(receipt, session)


@router.post(
    "/purchases/receipts",
    tags=["purchases"],
    status_code=status.HTTP_201_CREATED,
    response_model=PurchaseReceiptRead,
    dependencies=[Depends(require_permissions("purchases.create"))],
)
def create_purchase_receipt(
    payload: PurchaseReceiptCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> PurchaseReceiptRead:
    supplier = session.get(Supplier, payload.supplier_id)
    if not supplier or not supplier.is_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proveedor no encontrado")
    deposit = session.get(Deposit, payload.deposit_id)
    if not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Depósito no encontrado")
    if not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes cargar al menos un ítem")

    sku_ids = {item.sku_id for item in payload.items}
    skus = session.exec(select(SKU).where(SKU.id.in_(sku_ids))).all()
    if len(skus) != len(sku_ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Algún SKU no existe")

    audit_context = _build_audit_context(request)
    for item in payload.items:
        if item.quantity <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La cantidad debe ser mayor a cero")
        if deposit.controls_lot and not item.lot_code:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote es obligatorio para el depósito seleccionado")

    receipt = PurchaseReceipt(
        supplier_id=supplier.id,
        deposit_id=deposit.id,
        received_at=payload.received_at or date.today(),
        document_number=payload.document_number,
        notes=payload.notes,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
    )
    session.add(receipt)
    session.flush()

    movement_type = _get_movement_type_by_code(session, "PURCHASE")
    reference_value = payload.document_number or f"COMPRA-{receipt.id}"

    for item in payload.items:
        record = PurchaseReceiptItem(
            receipt_id=receipt.id,
            sku_id=item.sku_id,
            quantity=item.quantity,
            unit=item.unit,
            lot_code=item.lot_code,
            expiry_date=item.expiry_date,
            unit_cost=item.unit_cost,
        )
        session.add(record)
        session.flush()

        movement_payload = StockMovementCreate(
            sku_id=item.sku_id,
            deposit_id=deposit.id,
            movement_type_id=movement_type.id,
            quantity=item.quantity,
            unit=item.unit,
            lot_code=item.lot_code,
            expiry_date=item.expiry_date,
            reference_type="PURCHASE_RECEIPT",
            reference_id=receipt.id,
            reference_item_id=record.id,
            reference=reference_value,
            movement_date=receipt.received_at,
            created_by_user_id=current_user.id,
        )
        _, movement = _apply_stock_movement(
            session,
            movement_payload,
            allow_negative_balance=True,
            audit_context=audit_context,
            audit_movement=True,
        )
        record.stock_movement_id = movement.id
        record.updated_at = datetime.utcnow()
        session.add(record)

    _log_audit(
        session,
        "purchase_receipts",
        receipt.id,
        AuditAction.CREATE,
        current_user.id,
        _build_audit_changes(
            "create",
            "purchase_receipts",
            receipt.id,
            diff=_create_diff_from_fields(
                {
                    "supplier_id": supplier.id,
                    "deposit_id": deposit.id,
                    "received_at": receipt.received_at,
                    "document_number": receipt.document_number,
                    "notes": receipt.notes,
                    "items": len(payload.items),
                },
                ["supplier_id", "deposit_id", "received_at", "document_number", "notes", "items"],
            ),
            metadata=audit_context.get("metadata", {}),
        ),
        audit_context=audit_context,
    )
    session.commit()
    session.refresh(receipt)
    return _map_purchase_receipt(receipt, session)


def _ensure_stock_level(session: Session, deposit_id: int, sku_id: int) -> StockLevel:
    stock_level = session.exec(
        select(StockLevel).where(StockLevel.deposit_id == deposit_id, StockLevel.sku_id == sku_id)
    ).first()
    if stock_level:
        return stock_level

    stock_level = StockLevel(deposit_id=deposit_id, sku_id=sku_id, quantity=0)
    session.add(stock_level)
    session.flush()
    return stock_level


def _apply_stock_movement(
    session: Session,
    payload: StockMovementCreate,
    allow_negative_balance: bool = True,
    audit_context: dict | None = None,
    audit_movement: bool = False,
) -> tuple[StockLevel, StockMovement]:
    if payload.quantity <= 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La cantidad debe ser mayor a cero")

    movement_type = _get_movement_type_or_404(session, payload.movement_type_id)
    if not movement_type.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El tipo de movimiento está inactivo")

    sku = session.get(SKU, payload.sku_id)
    deposit = session.get(Deposit, payload.deposit_id)
    if not sku or not deposit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU o depósito no encontrado")
    session.refresh(sku, attribute_names=["sku_type"])
    if not (sku.sku_type and sku.sku_type.is_active):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El tipo del SKU está inactivo")
    production_line = None
    if payload.production_line_id:
        production_line = _get_production_line_or_404(session, payload.production_line_id)

    input_unit = payload.unit or sku.unit
    base_quantity = _convert_to_base_quantity(sku, payload.quantity, input_unit, session)
    movement_code = movement_type.code.upper()
    is_outgoing = payload.is_outgoing
    if is_outgoing is None:
        if movement_code in OUTGOING_MOVEMENTS:
            is_outgoing = True
        elif movement_code in INCOMING_MOVEMENTS:
            is_outgoing = False
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de movimiento no soportado")
    delta = -base_quantity if is_outgoing else base_quantity

    produced_at = payload.movement_date or date.today()
    if movement_code == "PRODUCTION" and not production_line:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La línea de producción es obligatoria")
    if movement_code == "PURCHASE" and deposit.controls_lot and not (payload.lot_code or payload.production_lot_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote es obligatorio para ingresos de compra")

    lot_code = payload.lot_code
    production_lot: ProductionLot | None = None
    created_lot = False
    if payload.production_lot_id:
        production_lot = session.get(ProductionLot, payload.production_lot_id)
        if not production_lot:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote de producción no encontrado")
        if production_lot.sku_id != sku.id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote no corresponde al SKU indicado")
        if production_lot.deposit_id != deposit.id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote pertenece a otro depósito")
        if production_lot.is_blocked:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote está bloqueado para movimientos")
        if payload.production_line_id and production_lot.production_line_id not in (None, payload.production_line_id):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote está asociado a otra línea")
        if movement_code == "PRODUCTION":
            if produced_at != production_lot.produced_at:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La fecha de producción no coincide con el lote")
            if not production_lot.production_line_id and production_line:
                production_lot.production_line_id = production_line.id
        lot_code = production_lot.lot_code
    elif lot_code:
        production_lot = session.exec(select(ProductionLot).where(ProductionLot.lot_code == lot_code)).first()
        if production_lot:
            if production_lot.sku_id != sku.id or production_lot.deposit_id != deposit.id:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote pertenece a otro SKU o depósito")
            if production_lot.is_blocked:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote está bloqueado para movimientos")
            if payload.production_line_id and production_lot.production_line_id not in (None, payload.production_line_id):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote está asociado a otra línea")
            validation_line = production_line or (
                production_lot.production_line_id and _get_production_line_or_404(session, production_lot.production_line_id)
            )
            if validation_line:
                _validate_lot_code(
                    session,
                    lot_code,
                    sku,
                    deposit,
                    validation_line,
                    production_lot.produced_at,
                    allow_existing_id=production_lot.id,
                )
        elif movement_code not in {"PRODUCTION", "PURCHASE"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote indicado no existe")

    if movement_code == "PRODUCTION" and not production_lot:
        if not production_line:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La línea de producción es obligatoria")
        lot_code = lot_code or _generate_lot_code(session, sku, production_line, produced_at)
        _validate_lot_code(session, lot_code, sku, deposit, production_line, produced_at)
        production_lot = ProductionLot(
            lot_code=lot_code,
            sku_id=sku.id,
            deposit_id=deposit.id,
            production_line_id=payload.production_line_id,
            produced_quantity=base_quantity,
            remaining_quantity=base_quantity,
            produced_at=produced_at,
            expiry_date=payload.expiry_date,
        )
        session.add(production_lot)
        session.flush()
        created_lot = True
    elif movement_code == "PRODUCTION" and production_lot:
        if produced_at != production_lot.produced_at:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La fecha de producción no coincide con el lote")
        validation_line = production_line or (
            production_lot.production_line_id and _get_production_line_or_404(session, production_lot.production_line_id)
        )
        if not validation_line:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote no tiene línea asociada")
        _validate_lot_code(
            session,
            production_lot.lot_code,
            sku,
            deposit,
            validation_line,
            production_lot.produced_at,
            allow_existing_id=production_lot.id,
        )
        if not production_lot.production_line_id and production_line:
            production_lot.production_line_id = production_line.id
        if payload.expiry_date and production_lot.expiry_date and payload.expiry_date != production_lot.expiry_date:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El vencimiento no coincide con el lote")
        if payload.expiry_date and not production_lot.expiry_date:
            production_lot.expiry_date = payload.expiry_date
    elif movement_code == "PURCHASE" and lot_code and not production_lot:
        production_lot = ProductionLot(
            lot_code=lot_code,
            sku_id=sku.id,
            deposit_id=deposit.id,
            produced_quantity=base_quantity,
            remaining_quantity=base_quantity,
            produced_at=produced_at,
            expiry_date=payload.expiry_date,
        )
        session.add(production_lot)
        session.flush()
        created_lot = True
    elif movement_code == "PURCHASE" and production_lot:
        if payload.expiry_date and production_lot.expiry_date and payload.expiry_date != production_lot.expiry_date:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El vencimiento no coincide con el lote")
        if payload.expiry_date and not production_lot.expiry_date:
            production_lot.expiry_date = payload.expiry_date

    stock_level = _ensure_stock_level(session, payload.deposit_id, payload.sku_id)
    new_quantity = stock_level.quantity + delta
    if not allow_negative_balance and new_quantity < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Stock insuficiente en el depósito")

    if production_lot:
        if movement_code == "PRODUCTION" and not created_lot:
            production_lot.produced_quantity += base_quantity
            production_lot.remaining_quantity += base_quantity
        elif created_lot and movement_code in {"PRODUCTION", "PURCHASE"}:
            pass
        else:
            new_remaining = production_lot.remaining_quantity + delta
            if not allow_negative_balance and new_remaining < 0:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote no tiene stock suficiente")
            production_lot.remaining_quantity = new_remaining
        production_lot.updated_at = datetime.utcnow()
        session.add(production_lot)

    stock_level.quantity = new_quantity
    movement = StockMovement(
        sku_id=payload.sku_id,
        deposit_id=payload.deposit_id,
        movement_type_id=movement_type.id,
        quantity=delta,
        reference_type=payload.reference_type,
        reference_id=payload.reference_id,
        reference_item_id=payload.reference_item_id,
        reference=payload.reference,
        lot_code=lot_code,
        production_lot_id=production_lot.id if production_lot else None,
        movement_date=payload.movement_date or date.today(),
        created_by_user_id=payload.created_by_user_id,
    )
    session.add(stock_level)
    session.add(movement)
    session.flush()
    session.refresh(stock_level, attribute_names=["sku", "deposit"])
    session.refresh(movement, attribute_names=["sku", "deposit", "movement_type", "production_lot"])

    if audit_movement:
        diff = {
            "sku_id": {"after": movement.sku_id},
            "deposit_id": {"after": movement.deposit_id},
            "movement_type_id": {"after": movement.movement_type_id},
            "quantity": {"after": movement.quantity},
            "reference_type": {"after": movement.reference_type},
            "reference_id": {"after": movement.reference_id},
            "reference_item_id": {"after": movement.reference_item_id},
            "reference": {"after": movement.reference},
            "lot_code": {"after": movement.lot_code},
            "production_lot_id": {"after": movement.production_lot_id},
            "movement_date": {"after": movement.movement_date},
        }
        changes = _build_audit_changes(
            "create",
            "stock_movements",
            movement.id,
            diff=diff,
            refs=_build_audit_refs(movement.reference_type, movement.reference_id),
            metadata=(audit_context or {}).get("metadata", {}),
        )
        _log_audit(
            session,
            "stock_movements",
            movement.id,
            AuditAction.CREATE,
            payload.created_by_user_id,
            changes,
            audit_context=audit_context,
        )
    if movement_code == "PRODUCTION" and sku.sku_type and sku.sku_type.code in SKU_PRODUCTION_TYPES:
        _consume_recipe_components(
            session,
            sku,
            deposit,
            production_lot,
            base_quantity,
            payload.reference,
            movement.movement_date,
            payload.created_by_user_id,
        )
    return stock_level, movement


def _map_stock_movement(movement: StockMovement, session: Session) -> StockMovementRead:
    session.refresh(movement, attribute_names=["sku", "deposit", "movement_type", "production_lot"])
    production_line_id = None
    production_line_name = None
    expiry_date = None
    if movement.production_lot:
        session.refresh(movement.production_lot, attribute_names=["production_line"])
        production_line_id = movement.production_lot.production_line_id
        production_line_name = (
            movement.production_lot.production_line.name if movement.production_lot.production_line else None
        )
        expiry_date = movement.production_lot.expiry_date
    stock_level = session.exec(
        select(StockLevel).where(StockLevel.deposit_id == movement.deposit_id, StockLevel.sku_id == movement.sku_id)
    ).first()
    current_balance = stock_level.quantity if stock_level else 0
    created_by_name = None
    if movement.created_by_user_id:
        user = session.get(User, movement.created_by_user_id)
        created_by_name = user.full_name if user else None

    return StockMovementRead(
        id=movement.id,
        sku_id=movement.sku_id,
        sku_code=movement.sku.code if movement.sku else str(movement.sku_id),
        sku_name=movement.sku.name if movement.sku else f"SKU {movement.sku_id}",
        deposit_id=movement.deposit_id,
        deposit_name=movement.deposit.name if movement.deposit else "",
        movement_type_id=movement.movement_type_id,
        movement_type_code=movement.movement_type.code if movement.movement_type else "",
        movement_type_label=movement.movement_type.label if movement.movement_type else "",
        quantity=movement.quantity,
        reference_type=movement.reference_type,
        reference_id=movement.reference_id,
        reference_item_id=movement.reference_item_id,
        reference=movement.reference,
        lot_code=movement.lot_code,
        production_lot_id=movement.production_lot_id,
        production_line_id=production_line_id,
        production_line_name=production_line_name,
        expiry_date=expiry_date,
        movement_date=movement.movement_date,
        created_at=movement.created_at,
        current_balance=current_balance,
        created_by_user_id=movement.created_by_user_id,
        created_by_name=created_by_name,
    )


def _map_inventory_count_item(item: InventoryCountItem, session: Session) -> InventoryCountItemRead:
    sku = session.get(SKU, item.sku_id)
    lot_code = item.lot_code
    if item.production_lot_id:
        lot = session.get(ProductionLot, item.production_lot_id)
        lot_code = lot.lot_code if lot else lot_code
    return InventoryCountItemRead(
        id=item.id,
        sku_id=item.sku_id,
        sku_code=sku.code if sku else str(item.sku_id),
        sku_name=sku.name if sku else f"SKU {item.sku_id}",
        production_lot_id=item.production_lot_id,
        lot_code=lot_code,
        counted_quantity=item.counted_quantity,
        system_quantity=item.system_quantity,
        difference=item.difference,
        unit=item.unit,
        stock_movement_id=item.stock_movement_id,
    )


def _map_inventory_count(count: InventoryCount, session: Session) -> InventoryCountRead:
    session.refresh(count, attribute_names=["items", "deposit"])
    created_by_name = None
    updated_by_name = None
    approved_by_name = None
    if count.created_by_user_id:
        user = session.get(User, count.created_by_user_id)
        created_by_name = user.full_name if user else None
    if count.updated_by_user_id:
        user = session.get(User, count.updated_by_user_id)
        updated_by_name = user.full_name if user else None
    if count.approved_by_user_id:
        user = session.get(User, count.approved_by_user_id)
        approved_by_name = user.full_name if user else None

    return InventoryCountRead(
        id=count.id,
        deposit_id=count.deposit_id,
        deposit_name=count.deposit.name if count.deposit else "",
        status=count.status,
        count_date=count.count_date,
        notes=count.notes,
        submitted_at=count.submitted_at,
        approved_at=count.approved_at,
        closed_at=count.closed_at,
        cancelled_at=count.cancelled_at,
        created_at=count.created_at,
        created_by_user_id=count.created_by_user_id,
        created_by_name=created_by_name,
        updated_by_user_id=count.updated_by_user_id,
        updated_by_name=updated_by_name,
        approved_by_user_id=count.approved_by_user_id,
        approved_by_name=approved_by_name,
        items=[_map_inventory_count_item(item, session) for item in count.items],
    )


def _map_supplier(record: Supplier) -> SupplierRead:
    return SupplierRead.model_validate(record)


def _map_purchase_receipt_item(item: PurchaseReceiptItem, session: Session) -> PurchaseReceiptItemRead:
    sku = session.get(SKU, item.sku_id)
    return PurchaseReceiptItemRead(
        id=item.id,
        sku_id=item.sku_id,
        sku_code=sku.code if sku else str(item.sku_id),
        sku_name=sku.name if sku else f"SKU {item.sku_id}",
        quantity=item.quantity,
        unit=item.unit,
        lot_code=item.lot_code,
        expiry_date=item.expiry_date,
        unit_cost=item.unit_cost,
        stock_movement_id=item.stock_movement_id,
    )


def _map_purchase_receipt(receipt: PurchaseReceipt, session: Session) -> PurchaseReceiptRead:
    session.refresh(receipt, attribute_names=["supplier", "deposit", "items"])
    created_by_name = None
    if receipt.created_by_user_id:
        user = session.get(User, receipt.created_by_user_id)
        created_by_name = user.full_name if user else None
    return PurchaseReceiptRead(
        id=receipt.id,
        supplier_id=receipt.supplier_id,
        supplier_name=receipt.supplier.name if receipt.supplier else None,
        deposit_id=receipt.deposit_id,
        deposit_name=receipt.deposit.name if receipt.deposit else None,
        received_at=receipt.received_at,
        document_number=receipt.document_number,
        notes=receipt.notes,
        created_at=receipt.created_at,
        created_by_user_id=receipt.created_by_user_id,
        created_by_name=created_by_name,
        items=[_map_purchase_receipt_item(item, session) for item in receipt.items],
    )


def _map_audit_log(record: AuditLog, session: Session) -> AuditLogRead:
    user_name = None
    if record.user_id:
        user = session.get(User, record.user_id)
        user_name = user.full_name if user else None
    return AuditLogRead(
        id=record.id,
        entity_type=record.entity_type,
        entity_id=record.entity_id,
        action=record.action,
        changes=record.changes,
        user_id=record.user_id,
        user_name=user_name,
        ip_address=record.ip_address,
        created_at=record.created_at,
    )


def _build_audit_logs_statement(
    entity_type: str | None = None,
    entity_id: int | None = None,
    user_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    detail_q: str | None = None,
) -> Select:
    statement = select(AuditLog)
    if entity_type:
        statement = statement.where(AuditLog.entity_type == entity_type)
    if entity_id is not None:
        statement = statement.where(AuditLog.entity_id == entity_id)
    if user_id:
        statement = statement.where(AuditLog.user_id == user_id)
    if date_from:
        statement = statement.where(AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        statement = statement.where(AuditLog.created_at <= datetime.combine(date_to, datetime.max.time()))
    if detail_q:
        statement = statement.where(cast(AuditLog.changes, String).ilike(f"%{detail_q}%"))
    return statement


def _summarize_audit_log(record: AuditLog, user: User | None, changes: dict | None) -> str:
    action_labels = {
        AuditAction.CREATE: "Alta",
        AuditAction.UPDATE: "Edición",
        AuditAction.DELETE: "Borrado",
        AuditAction.STATUS: "Cambio de estado",
        AuditAction.APPROVE: "Aprobación",
        AuditAction.CANCEL: "Cancelación",
    }
    action_label = action_labels.get(record.action, str(record.action))
    entity_part = record.entity_type
    entity_id = f"#{record.entity_id}" if record.entity_id is not None else ""
    event = ""
    if isinstance(changes, dict):
        event = str(changes.get("event") or "")
    user_part = ""
    if user:
        user_part = user.full_name or user.email
    summary_parts = [part for part in [action_label, event, entity_part, entity_id, user_part] if part]
    return " ".join(summary_parts)


def _map_merma_event(event: MermaEvent, session: Session) -> MermaEventRead:
    session.refresh(
        event,
        attribute_names=[
            "sku",
            "deposit",
            "production_line",
            "type",
            "cause",
            "remito",
            "order",
            "stock_movement",
            "reported_by_user",
        ],
    )
    return MermaEventRead(
        id=event.id,
        stage=event.stage,
        type_id=event.type_id,
        type_code=event.type_code,
        type_label=event.type_label,
        cause_id=event.cause_id,
        cause_code=event.cause_code,
        cause_label=event.cause_label,
        sku_id=event.sku_id,
        sku_code=event.sku.code if event.sku else str(event.sku_id),
        sku_name=event.sku.name if event.sku else f"SKU {event.sku_id}",
        quantity=event.quantity,
        unit=event.unit,
        lot_code=event.lot_code,
        deposit_id=event.deposit_id,
        deposit_name=event.deposit.name if event.deposit else None,
        remito_id=event.remito_id,
        order_id=event.order_id,
        production_line_id=event.production_line_id,
        production_line_name=event.production_line.name if event.production_line else None,
        reported_by_user_id=event.reported_by_user_id,
        reported_by_role=event.reported_by_role,
        notes=event.notes,
        detected_at=event.detected_at,
        created_at=event.created_at,
        updated_at=event.updated_at,
        affects_stock=event.affects_stock,
        action=event.action,
        stock_movement_id=event.stock_movement_id,
    )


@router.get(
    "/stock-levels",
    tags=["stock"],
    response_model=list[StockLevelRead],
    dependencies=[Depends(require_permissions("stock.view"))],
)
def list_stock_levels(session: Session = Depends(get_session)) -> list[StockLevelRead]:
    stock_levels = session.exec(select(StockLevel)).all()
    return [_map_stock_level(level, session) for level in stock_levels]


@router.post(
    "/stock/movements",
    tags=["stock"],
    status_code=status.HTTP_201_CREATED,
    response_model=StockLevelRead,
    dependencies=[Depends(require_permissions("stock.register"))],
)
def register_stock_movement(
    payload: StockMovementCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> StockLevelRead:
    payload.created_by_user_id = payload.created_by_user_id or current_user.id
    audit_context = _build_audit_context(request)
    stock_level, _ = _apply_stock_movement(session, payload, audit_context=audit_context, audit_movement=True)
    session.commit()
    session.refresh(stock_level, attribute_names=["sku", "deposit"])
    return _map_stock_level(stock_level, session)


@router.get(
    "/stock/movements",
    tags=["stock"],
    response_model=StockMovementList,
    dependencies=[Depends(require_permissions("stock.view"))],
)
def list_stock_movements(
    sku_id: int | None = None,
    deposit_id: int | None = None,
    movement_type_id: int | None = None,
    movement_type_code: str | None = None,
    production_line_id: int | None = None,
    lot_code: str | None = None,
    reference_type: str | None = None,
    reference_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    limit: int = 50,
    offset: int = 0,
    session: Session = Depends(get_session),
) -> StockMovementList:
    statement = select(StockMovement)
    if sku_id:
        statement = statement.where(StockMovement.sku_id == sku_id)
    if deposit_id:
        statement = statement.where(StockMovement.deposit_id == deposit_id)
    if movement_type_id:
        statement = statement.where(StockMovement.movement_type_id == movement_type_id)
    if movement_type_code:
        statement = statement.join(StockMovementType).where(
            StockMovementType.code == movement_type_code.strip().upper()
        )
    if production_line_id:
        statement = statement.join(ProductionLot).where(ProductionLot.production_line_id == production_line_id)
    if lot_code:
        statement = statement.where(StockMovement.lot_code == lot_code)
    if reference_type:
        statement = statement.where(func.upper(StockMovement.reference_type) == reference_type.strip().upper())
    if reference_id is not None:
        statement = statement.where(StockMovement.reference_id == reference_id)
    if date_from:
        statement = statement.where(StockMovement.movement_date >= date_from)
    if date_to:
        statement = statement.where(StockMovement.movement_date <= date_to)

    safe_limit = max(1, min(limit, 200))
    safe_offset = max(offset, 0)
    result = session.exec(
        select(func.count()).select_from(statement.subquery())
    )
    total = result.first()
    if isinstance(total, tuple):
        total = total[0]

    records = session.exec(
        statement.order_by(StockMovement.movement_date.desc(), StockMovement.id.desc())
        .offset(safe_offset)
        .limit(safe_limit)
    ).all()
    return StockMovementList(total=total or 0, items=[_map_stock_movement(item, session) for item in records])


def _resolve_system_quantity(
    session: Session,
    deposit: Deposit,
    sku: SKU,
    production_lot_id: int | None,
) -> tuple[float, ProductionLot | None]:
    if deposit.controls_lot:
        if not production_lot_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El depósito requiere lote")
        lot = session.get(ProductionLot, production_lot_id)
        if not lot:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote no encontrado")
        if lot.deposit_id != deposit.id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote pertenece a otro depósito")
        if lot.sku_id != sku.id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El lote no corresponde al SKU indicado")
        return float(lot.remaining_quantity), lot

    stock_level = session.exec(
        select(StockLevel).where(StockLevel.deposit_id == deposit.id, StockLevel.sku_id == sku.id)
    ).first()
    return float(stock_level.quantity) if stock_level else 0.0, None


def _replace_inventory_count_items(
    session: Session,
    count: InventoryCount,
    items: list[dict],
    deposit: Deposit,
) -> None:
    existing_items = session.exec(select(InventoryCountItem).where(InventoryCountItem.inventory_count_id == count.id)).all()
    for item in existing_items:
        session.delete(item)
    session.flush()

    for payload in items:
        sku = session.get(SKU, payload["sku_id"])
        if not sku:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU no encontrado")
        system_quantity, lot = _resolve_system_quantity(session, deposit, sku, payload.get("production_lot_id"))
        counted = float(payload["counted_quantity"])
        difference = counted - system_quantity
        session.add(
            InventoryCountItem(
                inventory_count_id=count.id,
                sku_id=sku.id,
                production_lot_id=payload.get("production_lot_id"),
                lot_code=lot.lot_code if lot else None,
                counted_quantity=counted,
                system_quantity=system_quantity,
                difference=difference,
                unit=sku.unit,
            )
        )


def _serialize_inventory_count_items(items: list[InventoryCountItem]) -> list[dict]:
    return [
        {
            "sku_id": item.sku_id,
            "production_lot_id": item.production_lot_id,
            "counted_quantity": item.counted_quantity,
            "system_quantity": item.system_quantity,
            "difference": item.difference,
        }
        for item in items
    ]


@router.get(
    "/inventory-counts",
    tags=["inventory"],
    response_model=list[InventoryCountRead],
    dependencies=[Depends(require_permissions("inventory.view"))],
)
def list_inventory_counts(
    status_filter: InventoryCountStatus | None = None,
    deposit_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    session: Session = Depends(get_session),
) -> list[InventoryCountRead]:
    statement = select(InventoryCount)
    if status_filter:
        statement = statement.where(InventoryCount.status == status_filter)
    if deposit_id:
        statement = statement.where(InventoryCount.deposit_id == deposit_id)
    if date_from:
        statement = statement.where(InventoryCount.count_date >= date_from)
    if date_to:
        statement = statement.where(InventoryCount.count_date <= date_to)
    counts = session.exec(statement.order_by(InventoryCount.count_date.desc(), InventoryCount.id.desc())).all()
    return [_map_inventory_count(count, session) for count in counts]


@router.get(
    "/inventory-counts/{count_id}",
    tags=["inventory"],
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.view"))],
)
def get_inventory_count(count_id: int, session: Session = Depends(get_session)) -> InventoryCountRead:
    count = session.get(InventoryCount, count_id)
    if not count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conteo no encontrado")
    return _map_inventory_count(count, session)


@router.post(
    "/inventory-counts",
    tags=["inventory"],
    status_code=status.HTTP_201_CREATED,
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.create"))],
)
def create_inventory_count(
    payload: InventoryCountCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> InventoryCountRead:
    if not payload.items:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debes cargar al menos un ítem")
    deposit = _get_deposit_or_404(session, payload.deposit_id)

    count = InventoryCount(
        deposit_id=deposit.id,
        status=InventoryCountStatus.DRAFT,
        count_date=payload.count_date or date.today(),
        notes=payload.notes,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
    )
    session.add(count)
    session.flush()

    items_payload = [item.model_dump() for item in payload.items]
    _replace_inventory_count_items(session, count, items_payload, deposit)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "create",
        "inventory_counts",
        count.id,
        diff=_create_diff_from_fields(
            {
                "deposit_id": count.deposit_id,
                "count_date": count.count_date,
                "notes": count.notes,
                "status": count.status,
                "items": items_payload,
            },
            ["deposit_id", "count_date", "notes", "status", "items"],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "inventory_counts", count.id, AuditAction.CREATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(count)
    return _map_inventory_count(count, session)


@router.put(
    "/inventory-counts/{count_id}",
    tags=["inventory"],
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.edit"))],
)
def update_inventory_count(
    count_id: int,
    payload: InventoryCountUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> InventoryCountRead:
    count = session.get(InventoryCount, count_id)
    if not count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conteo no encontrado")
    if count.status != InventoryCountStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden editar conteos en borrador")

    session.refresh(count, attribute_names=["items"])
    before_items = [
        {
            "sku_id": item.sku_id,
            "production_lot_id": item.production_lot_id,
            "counted_quantity": item.counted_quantity,
        }
        for item in count.items
    ]
    update_data = payload.model_dump(exclude_unset=True)
    items = update_data.pop("items", None)
    items_payload = None
    if items is not None:
        items_payload = [item.model_dump() if hasattr(item, "model_dump") else item for item in items]
    diff = _build_diff_from_payload(count, update_data)
    if items_payload is not None and before_items != items_payload:
        diff["items"] = {"before": before_items, "after": items_payload}
    for field, value in update_data.items():
        setattr(count, field, value)
    count.updated_at = datetime.utcnow()
    count.updated_by_user_id = current_user.id
    session.add(count)
    session.flush()

    if items_payload is not None:
        deposit = _get_deposit_or_404(session, count.deposit_id)
        _replace_inventory_count_items(session, count, items_payload, deposit)

    audit_context = _build_audit_context(request)
    if diff:
        changes = _build_audit_changes(
            "update",
            "inventory_counts",
            count.id,
            diff=diff,
            metadata=audit_context.get("metadata", {}),
        )
        _log_audit(session, "inventory_counts", count.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(count)
    return _map_inventory_count(count, session)


@router.post(
    "/inventory-counts/{count_id}/submit",
    tags=["inventory"],
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.close"))],
)
def submit_inventory_count(
    count_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> InventoryCountRead:
    count = session.get(InventoryCount, count_id)
    if not count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conteo no encontrado")
    if count.status != InventoryCountStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El conteo ya fue enviado")

    count.status = InventoryCountStatus.SUBMITTED
    count.submitted_at = datetime.utcnow()
    count.updated_at = datetime.utcnow()
    count.updated_by_user_id = current_user.id
    session.add(count)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "status",
        "inventory_counts",
        count.id,
        diff=_status_diff("status", InventoryCountStatus.DRAFT, count.status),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "inventory_counts", count.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(count)
    return _map_inventory_count(count, session)


@router.post(
    "/inventory-counts/{count_id}/approve",
    tags=["inventory"],
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.close"))],
)
def approve_inventory_count(
    count_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> InventoryCountRead:
    count = session.get(InventoryCount, count_id)
    if not count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conteo no encontrado")
    if count.status != InventoryCountStatus.SUBMITTED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El conteo debe estar enviado para aprobarse")

    session.refresh(count, attribute_names=["items"])
    deposit = _get_deposit_or_404(session, count.deposit_id)
    movement_type = _get_movement_type_by_code(session, "ADJUSTMENT")
    reference = f"INVENTARIO-{count.id}"

    audit_context = _build_audit_context(request)
    for item in count.items:
        if item.difference == 0:
            continue
        if deposit.controls_lot and not item.production_lot_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Falta lote en depósito con control de lotes")

        lot_code = item.lot_code
        if item.production_lot_id:
            lot = session.get(ProductionLot, item.production_lot_id)
            lot_code = lot.lot_code if lot else lot_code

        movement_payload = StockMovementCreate(
            sku_id=item.sku_id,
            deposit_id=deposit.id,
            movement_type_id=movement_type.id,
            quantity=abs(item.difference),
            is_outgoing=item.difference < 0,
            reference_type="INVENTORY_COUNT",
            reference_id=count.id,
            reference_item_id=item.id,
            reference=reference,
            lot_code=lot_code,
            production_lot_id=item.production_lot_id,
            movement_date=date.today(),
            created_by_user_id=current_user.id,
        )
        _, movement = _apply_stock_movement(
            session,
            movement_payload,
            allow_negative_balance=False,
            audit_context=audit_context,
            audit_movement=True,
        )
        item.stock_movement_id = movement.id
        item.updated_at = datetime.utcnow()
        session.add(item)

    count.status = InventoryCountStatus.APPROVED
    count.approved_at = datetime.utcnow()
    count.approved_by_user_id = current_user.id
    count.updated_at = datetime.utcnow()
    count.updated_by_user_id = current_user.id
    session.add(count)
    changes = _build_audit_changes(
        "approve",
        "inventory_counts",
        count.id,
        diff=_status_diff("status", InventoryCountStatus.SUBMITTED, count.status),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "inventory_counts", count.id, AuditAction.APPROVE, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(count)
    return _map_inventory_count(count, session)


@router.post(
    "/inventory-counts/{count_id}/close",
    tags=["inventory"],
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.close"))],
)
def close_inventory_count(
    count_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> InventoryCountRead:
    count = session.get(InventoryCount, count_id)
    if not count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conteo no encontrado")
    if count.status != InventoryCountStatus.APPROVED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Solo se pueden cerrar conteos aprobados")

    count.status = InventoryCountStatus.CLOSED
    count.closed_at = datetime.utcnow()
    count.updated_at = datetime.utcnow()
    count.updated_by_user_id = current_user.id
    session.add(count)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "status",
        "inventory_counts",
        count.id,
        diff=_status_diff("status", InventoryCountStatus.APPROVED, count.status),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "inventory_counts", count.id, AuditAction.STATUS, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(count)
    return _map_inventory_count(count, session)


@router.post(
    "/inventory-counts/{count_id}/cancel",
    tags=["inventory"],
    response_model=InventoryCountRead,
    dependencies=[Depends(require_permissions("inventory.close"))],
)
def cancel_inventory_count(
    count_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> InventoryCountRead:
    count = session.get(InventoryCount, count_id)
    if not count:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Conteo no encontrado")
    if count.status not in {InventoryCountStatus.DRAFT, InventoryCountStatus.SUBMITTED}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede cancelar el conteo")

    before_status = count.status
    count.status = InventoryCountStatus.CANCELLED
    count.cancelled_at = datetime.utcnow()
    count.updated_at = datetime.utcnow()
    count.updated_by_user_id = current_user.id
    session.add(count)
    audit_context = _build_audit_context(request)
    changes = _build_audit_changes(
        "cancel",
        "inventory_counts",
        count.id,
        diff=_status_diff("status", before_status, InventoryCountStatus.CANCELLED),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "inventory_counts", count.id, AuditAction.CANCEL, current_user.id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(count)
    return _map_inventory_count(count, session)


@router.get(
    "/audit/logs",
    tags=["audit"],
    response_model=list[AuditLogRead],
    dependencies=[Depends(require_permissions("audit.view"))],
)
def list_audit_logs(
    entity_type: str | None = None,
    entity_id: int | None = None,
    user_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    detail_q: str | None = None,
    limit: int = 200,
    session: Session = Depends(get_session),
) -> list[AuditLogRead]:
    statement = _build_audit_logs_statement(entity_type, entity_id, user_id, date_from, date_to, detail_q)
    safe_limit = max(1, min(limit, 500))
    records = session.exec(statement.order_by(AuditLog.created_at.desc(), AuditLog.id.desc()).limit(safe_limit)).all()
    return [_map_audit_log(record, session) for record in records]


def _build_excel_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _excel_safe_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool, datetime, date)):
        return value
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    return str(value)


@router.get(
    "/audit/logs/export.xlsx",
    tags=["audit"],
    dependencies=[Depends(require_permissions("audit.view"))],
)
def export_audit_logs(
    entity_type: str | None = None,
    entity_id: int | None = None,
    user_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    detail_q: str | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    statement = _build_audit_logs_statement(entity_type, entity_id, user_id, date_from, date_to, detail_q)
    records = session.exec(
        statement.order_by(AuditLog.created_at.desc(), AuditLog.id.desc()).limit(AUDIT_EXPORT_LIMIT + 1)
    ).all()
    if len(records) > AUDIT_EXPORT_LIMIT:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Demasiados registros para exportar; acotá filtros.",
        )
    user_ids = {record.user_id for record in records if record.user_id}
    users = session.exec(select(User).where(User.id.in_(user_ids))).all() if user_ids else []
    user_map = {user.id: user for user in users}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AuditLogs"
    headers = [
        "created_at",
        "user",
        "user_id",
        "action",
        "entity_type",
        "entity_id",
        "ip",
        "request_id",
        "summary",
        "detail",
    ]
    sheet.append(headers)
    changes_rows: list[tuple] = []
    for record in records:
        user = user_map.get(record.user_id) if record.user_id else None
        user_label = user.full_name or user.email if user else ""
        changes = record.changes if isinstance(record.changes, dict) else {}
        metadata = changes.get("metadata") if isinstance(changes, dict) else {}
        metadata = metadata if isinstance(metadata, dict) else {}
        ip_address = record.ip_address or metadata.get("ip")
        request_id = metadata.get("request_id")
        summary = _summarize_audit_log(record, user, changes)
        detail = json.dumps(changes, ensure_ascii=False, separators=(",", ":"), default=str) if changes else ""
        sheet.append(
            [
                record.created_at.isoformat(),
                user_label,
                record.user_id,
                record.action.value if isinstance(record.action, AuditAction) else str(record.action),
                record.entity_type,
                record.entity_id,
                ip_address,
                request_id,
                summary,
                detail,
            ]
        )
        diff = changes.get("diff") if isinstance(changes, dict) else None
        if isinstance(diff, dict):
            for field, values in diff.items():
                if not isinstance(values, dict):
                    continue
                changes_rows.append(
                    (
                        record.id,
                        record.created_at.isoformat(),
                        record.entity_type,
                        record.entity_id,
                        record.action.value if isinstance(record.action, AuditAction) else str(record.action),
                        user_label,
                        field,
                        _excel_safe_value(values.get("before")),
                        _excel_safe_value(values.get("after")),
                    )
                )
    if changes_rows:
        changes_sheet = workbook.create_sheet(title="ChangesFlatten")
        changes_sheet.append(
            [
                "audit_log_id",
                "created_at",
                "entity_type",
                "entity_id",
                "action",
                "user",
                "field",
                "before",
                "after",
            ]
        )
        for row in changes_rows:
            changes_sheet.append(list(row))
    filename = f"audit_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return _build_excel_response(workbook, filename)


@router.get(
    "/production/lots",
    tags=["production"],
    response_model=list[ProductionLotRead],
    dependencies=[Depends(require_permissions("production.view"))],
)
def list_production_lots(
    deposit_id: int | None = None,
    sku_id: int | None = None,
    production_line_id: int | None = None,
    lot_code: str | None = None,
    available_only: bool = False,
    include_blocked: bool = False,
    session: Session = Depends(get_session),
) -> list[ProductionLotRead]:
    statement = select(ProductionLot)
    if deposit_id:
        statement = statement.where(ProductionLot.deposit_id == deposit_id)
    if sku_id:
        statement = statement.where(ProductionLot.sku_id == sku_id)
    if production_line_id:
        statement = statement.where(ProductionLot.production_line_id == production_line_id)
    if lot_code:
        statement = statement.where(ProductionLot.lot_code == lot_code)
    if available_only:
        statement = statement.where(ProductionLot.remaining_quantity > 0)
    if not include_blocked:
        statement = statement.where(ProductionLot.is_blocked.is_(False))
    lots = session.exec(statement.order_by(ProductionLot.produced_at.desc(), ProductionLot.id.desc())).all()
    return [_map_production_lot(lot, session) for lot in lots]


@router.get(
    "/production/lots/{lot_id}",
    tags=["production"],
    response_model=ProductionLotRead,
    dependencies=[Depends(require_permissions("production.view"))],
)
def get_production_lot(lot_id: int, session: Session = Depends(get_session)) -> ProductionLotRead:
    lot = session.get(ProductionLot, lot_id)
    if not lot:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote de producción no encontrado")
    return _map_production_lot(lot, session)


@router.post(
    "/mermas",
    tags=["mermas"],
    status_code=status.HTTP_201_CREATED,
    response_model=MermaEventRead,
    dependencies=[Depends(require_permissions("mermas.report"))],
)
def create_merma_event(
    payload: MermaEventCreate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> MermaEventRead:
    merma_type = session.get(MermaType, payload.type_id)
    cause = session.get(MermaCause, payload.cause_id)
    sku = session.get(SKU, payload.sku_id)

    if not merma_type or merma_type.stage != payload.stage:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de merma inválido para la etapa")
    if not cause or cause.stage != payload.stage:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Causa inválida para la etapa")
    if not sku:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="SKU no encontrado")
    if not merma_type.is_active or not cause.is_active:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El tipo o la causa están inactivos")
    if payload.quantity <= 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La cantidad debe ser mayor a cero")

    production_line = None
    deposit = None
    remito = None
    order = None

    detected_at = payload.detected_at or datetime.utcnow()
    unit = payload.unit or sku.unit

    if payload.stage == MermaStage.PRODUCTION:
        if not payload.production_line_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La línea de producción es obligatoria")
        production_line = session.get(ProductionLine, payload.production_line_id)
        if not production_line or not production_line.is_active:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Línea de producción inválida")
        if not payload.deposit_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El depósito es obligatorio")
        deposit = session.get(Deposit, payload.deposit_id)
    elif payload.stage in {MermaStage.EMPAQUE, MermaStage.STOCK}:
        if not payload.deposit_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El depósito es obligatorio")
        deposit = session.get(Deposit, payload.deposit_id)
    elif payload.stage == MermaStage.TRANSITO_POST_REMITO:
        if not payload.remito_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El remito es obligatorio")
        remito = session.get(Remito, payload.remito_id)
        if not remito:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Remito no encontrado")
        if remito.order_id:
            order = session.get(Order, remito.order_id)
        if remito.destination_deposit_id:
            deposit = session.get(Deposit, remito.destination_deposit_id)
        elif remito.shipment_id:
            shipment = session.get(Shipment, remito.shipment_id)
            if shipment:
                deposit = session.get(Deposit, shipment.deposit_id)
        if not deposit:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El remito no tiene destino asociado")
        if order:
            payload.order_id = order.id
        payload.deposit_id = deposit.id if deposit else payload.deposit_id
    elif payload.stage == MermaStage.ADMINISTRATIVA:
        if not payload.notes or not payload.notes.strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Notas obligatorias en ajustes administrativos")
        if payload.affects_stock and not payload.deposit_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecciona el depósito a ajustar")
        if payload.deposit_id:
            deposit = session.get(Deposit, payload.deposit_id)
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Etapa no soportada")

    if payload.deposit_id and not deposit and payload.stage != MermaStage.TRANSITO_POST_REMITO:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Depósito no encontrado")
    if payload.affects_stock and not deposit:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede ajustar stock sin depósito")
    reported_by_user_id = payload.reported_by_user_id or current_user.id
    reported_by_role = payload.reported_by_role
    if not reported_by_role and current_user.role_id:
        role = session.get(Role, current_user.role_id)
        reported_by_role = role.name if role else None

    if reported_by_user_id:
        user = session.get(User, reported_by_user_id)
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario informante no encontrado")

    audit_context = _build_audit_context(request)
    stock_movement_id = None
    if payload.affects_stock and deposit:
        merma_movement_type = _get_movement_type_by_code(session, "MERMA")
        movement_payload = StockMovementCreate(
            sku_id=payload.sku_id,
            deposit_id=deposit.id,
            movement_type_id=merma_movement_type.id,
            quantity=payload.quantity,
            unit=unit,
            reference=f"MERMA-{merma_type.code}",
            lot_code=payload.lot_code,
            movement_date=detected_at.date(),
            created_by_user_id=reported_by_user_id,
        )
        _, movement = _apply_stock_movement(
            session,
            movement_payload,
            allow_negative_balance=True,
            audit_context=audit_context,
            audit_movement=True,
        )
        stock_movement_id = movement.id

    event = MermaEvent(
        stage=payload.stage,
        type_id=merma_type.id,
        type_code=merma_type.code,
        type_label=merma_type.label,
        cause_id=cause.id,
        cause_code=cause.code,
        cause_label=cause.label,
        sku_id=payload.sku_id,
        quantity=payload.quantity,
        unit=unit,
        lot_code=payload.lot_code,
        deposit_id=payload.deposit_id,
        remito_id=payload.remito_id,
        order_id=payload.order_id,
        production_line_id=payload.production_line_id,
        reported_by_user_id=reported_by_user_id,
        reported_by_role=reported_by_role,
        notes=payload.notes,
        detected_at=detected_at,
        affects_stock=payload.affects_stock,
        action=payload.action,
        stock_movement_id=stock_movement_id,
    )

    session.add(event)
    session.flush()
    changes = _build_audit_changes(
        "create",
        "mermas",
        event.id,
        diff=_create_diff_from_fields(
            {
                "stage": event.stage,
                "type_id": event.type_id,
                "cause_id": event.cause_id,
                "sku_id": event.sku_id,
                "quantity": event.quantity,
                "unit": event.unit,
                "deposit_id": event.deposit_id,
                "remito_id": event.remito_id,
                "order_id": event.order_id,
                "production_line_id": event.production_line_id,
                "affects_stock": event.affects_stock,
                "action": event.action,
                "stock_movement_id": event.stock_movement_id,
                "reported_by_user_id": event.reported_by_user_id,
                "notes": event.notes,
            },
            [
                "stage",
                "type_id",
                "cause_id",
                "sku_id",
                "quantity",
                "unit",
                "deposit_id",
                "remito_id",
                "order_id",
                "production_line_id",
                "affects_stock",
                "action",
                "stock_movement_id",
                "reported_by_user_id",
                "notes",
            ],
        ),
        metadata=audit_context.get("metadata", {}),
    )
    _log_audit(session, "mermas", event.id, AuditAction.CREATE, reported_by_user_id, changes, audit_context=audit_context)
    session.commit()
    session.refresh(event)
    return _map_merma_event(event, session)


@router.get(
    "/mermas",
    tags=["mermas"],
    response_model=list[MermaEventRead],
    dependencies=[Depends(require_permissions("mermas.view"))],
)
def list_merma_events(
    date_from: date | None = None,
    date_to: date | None = None,
    stage: MermaStage | None = None,
    deposit_id: int | None = None,
    production_line_id: int | None = None,
    sku_id: int | None = None,
    type_id: int | None = None,
    cause_id: int | None = None,
    affects_stock: bool | None = None,
    session: Session = Depends(get_session),
) -> list[MermaEventRead]:
    statement = select(MermaEvent)
    if date_from:
        statement = statement.where(MermaEvent.detected_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        statement = statement.where(MermaEvent.detected_at <= datetime.combine(date_to, datetime.max.time()))
    if stage:
        statement = statement.where(MermaEvent.stage == stage)
    if deposit_id:
        statement = statement.where(MermaEvent.deposit_id == deposit_id)
    if production_line_id:
        statement = statement.where(MermaEvent.production_line_id == production_line_id)
    if sku_id:
        statement = statement.where(MermaEvent.sku_id == sku_id)
    if type_id:
        statement = statement.where(MermaEvent.type_id == type_id)
    if cause_id:
        statement = statement.where(MermaEvent.cause_id == cause_id)
    if affects_stock is not None:
        statement = statement.where(MermaEvent.affects_stock.is_(affects_stock))

    events = session.exec(statement.order_by(MermaEvent.detected_at.desc(), MermaEvent.id.desc())).all()
    return [_map_merma_event(event, session) for event in events]


@router.get(
    "/mermas/{merma_id}",
    tags=["mermas"],
    response_model=MermaEventRead,
    dependencies=[Depends(require_permissions("mermas.view"))],
)
def get_merma_event(merma_id: int, session: Session = Depends(get_session)) -> MermaEventRead:
    event = session.get(MermaEvent, merma_id)
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Merma no encontrada")
    return _map_merma_event(event, session)


@router.get(
    "/reports/stock-summary",
    tags=["reports"],
    response_model=StockReportRead,
    dependencies=[Depends(require_permissions("reports.view"))],
)
def stock_summary(session: Session = Depends(get_session)) -> StockReportRead:
    stock_levels = session.exec(select(StockLevel)).all()
    totals_by_tag: dict[str, float] = {}
    totals_by_deposit: dict[str, float] = {}

    for level in stock_levels:
        session.refresh(level, attribute_names=["sku", "deposit"])
        session.refresh(level.sku, attribute_names=["sku_type"])
        type_code = level.sku.sku_type.code if level.sku and level.sku.sku_type else "SIN_TIPO"
        totals_by_tag[type_code] = totals_by_tag.get(type_code, 0) + level.quantity
        totals_by_deposit[level.deposit.name] = totals_by_deposit.get(level.deposit.name, 0) + level.quantity

    movements_cutoff = date.today() - timedelta(days=7)
    movements = session.exec(select(StockMovement).where(StockMovement.movement_date >= movements_cutoff)).all()
    movement_totals: dict[str, dict[str, float | str]] = {}
    for mov in movements:
        session.refresh(mov, attribute_names=["movement_type"])
        code = mov.movement_type.code if mov.movement_type else "UNKNOWN"
        label = mov.movement_type.label if mov.movement_type else code
        if code not in movement_totals:
            movement_totals[code] = {"quantity": 0.0, "label": label}
        movement_totals[code]["quantity"] = float(movement_totals[code]["quantity"]) + mov.quantity

    return StockReportRead(
        totals_by_tag=[StockSummaryRow(group="tag", label=tag, quantity=qty) for tag, qty in totals_by_tag.items()],
        totals_by_deposit=[
          StockSummaryRow(group="deposit", label=deposit, quantity=qty) for deposit, qty in totals_by_deposit.items()
        ],
        movement_totals=[
            MovementSummary(movement_type_code=code, movement_type_label=data["label"], quantity=data["quantity"])
            for code, data in movement_totals.items()
        ],
    )


@router.get(
    "/reports/stock-alerts",
    tags=["reports"],
    response_model=StockAlertReport,
    dependencies=[Depends(require_permissions("reports.view"))],
)
def stock_alerts_report(
    sku_type_ids: list[int] | None = Query(None),
    deposit_ids: list[int] | None = Query(None),
    alert_status: list[str] | None = Query(None),
    search: str | None = None,
    min_quantity: float | None = None,
    max_quantity: float | None = None,
    only_configured: bool = False,
    include_inactive: bool = False,
    session: Session = Depends(get_session),
) -> StockAlertReport:
    statement = select(StockLevel).join(SKU).join(SKUType).join(Deposit)
    if sku_type_ids:
        statement = statement.where(SKU.sku_type_id.in_(sku_type_ids))
    if deposit_ids:
        statement = statement.where(StockLevel.deposit_id.in_(deposit_ids))
    if not include_inactive:
        statement = statement.where(SKU.is_active.is_(True), SKUType.is_active.is_(True))
    if search:
        like = f"%{search.lower()}%"
        statement = statement.where((SKU.name.ilike(like)) | (SKU.code.ilike(like)))
    if min_quantity is not None:
        statement = statement.where(StockLevel.quantity >= min_quantity)
    if max_quantity is not None:
        statement = statement.where(StockLevel.quantity <= max_quantity)

    levels = session.exec(statement.order_by(StockLevel.quantity.asc(), StockLevel.id.asc())).all()
    items: list[StockAlertRead] = []
    for level in levels:
        session.refresh(level, attribute_names=["sku", "deposit"])
        session.refresh(level.sku, attribute_names=["sku_type"])
        status = _get_stock_alert_status(level.quantity, level.sku)
        if only_configured and not _has_alert_thresholds(level.sku):
            continue
        if alert_status and status not in alert_status:
            continue
        items.append(
            StockAlertRead(
                deposit_id=level.deposit_id,
                deposit_name=level.deposit.name,
                sku_id=level.sku_id,
                sku_code=level.sku.code,
                sku_name=level.sku.name,
                sku_type_id=level.sku.sku_type_id,
                sku_type_code=level.sku.sku_type.code if level.sku.sku_type else "",
                sku_type_label=level.sku.sku_type.label if level.sku.sku_type else "",
                unit=level.sku.unit,
                quantity=level.quantity,
                alert_status=status,
                alert_green_min=level.sku.alert_green_min,
                alert_yellow_min=level.sku.alert_yellow_min,
            )
        )

    return StockAlertReport(total=len(items), items=items)


@router.get(
    "/reports/stock-expirations",
    tags=["reports"],
    response_model=ExpiryReport,
    dependencies=[Depends(require_permissions("reports.view"))],
)
def stock_expirations_report(
    sku_id: int | None = None,
    deposit_id: int | None = None,
    status: list[ExpiryReportStatus] | None = Query(None),
    expiry_from: date | None = None,
    expiry_to: date | None = None,
    include_no_expiry: bool = True,
    session: Session = Depends(get_session),
) -> ExpiryReport:
    statement = select(ProductionLot).join(SKU).join(Deposit)
    if sku_id:
        statement = statement.where(ProductionLot.sku_id == sku_id)
    if deposit_id:
        statement = statement.where(ProductionLot.deposit_id == deposit_id)
    expiry_predicates = []
    if expiry_from:
        expiry_predicates.append(ProductionLot.expiry_date >= expiry_from)
    if expiry_to:
        expiry_predicates.append(ProductionLot.expiry_date <= expiry_to)
    if expiry_predicates:
        expiry_clause = and_(*expiry_predicates)
        if include_no_expiry:
            statement = statement.where(or_(ProductionLot.expiry_date.is_(None), expiry_clause))
        else:
            statement = statement.where(expiry_clause)
    elif not include_no_expiry:
        statement = statement.where(ProductionLot.expiry_date.is_not(None))

    lots = session.exec(
        statement.where(ProductionLot.remaining_quantity > 0).order_by(
            nulls_last(ProductionLot.expiry_date),
            ProductionLot.produced_at,
            ProductionLot.id,
        )
    ).all()

    today = date.today()
    items: list[ExpiryReportRow] = []
    for lot in lots:
        session.refresh(lot, attribute_names=["sku", "deposit"])
        status_value, days = _get_expiry_status(lot.expiry_date, today)
        if status and status_value not in status:
            continue
        items.append(
            ExpiryReportRow(
                lot_id=lot.id,
                lot_code=lot.lot_code,
                sku_id=lot.sku_id,
                sku_code=lot.sku.code if lot.sku else str(lot.sku_id),
                sku_name=lot.sku.name if lot.sku else f"SKU {lot.sku_id}",
                deposit_id=lot.deposit_id,
                deposit_name=lot.deposit.name if lot.deposit else "",
                remaining_quantity=float(lot.remaining_quantity),
                unit=lot.sku.unit if lot.sku else UnitOfMeasure.UNIT,
                produced_at=lot.produced_at,
                expiry_date=lot.expiry_date,
                days_to_expiry=days,
                status=status_value,
                source_type="production",
                source_id=lot.id,
            )
        )

    return ExpiryReport(total=len(items), items=items)


@router.patch(
    "/lots/{source_type}/{source_id}",
    tags=["inventory"],
    response_model=LotExpiryUpdateResponse,
    dependencies=[Depends(require_permissions("production.edit", "purchases.edit"))],
)
def update_lot_expiry_date(
    source_type: str,
    source_id: int,
    payload: LotExpiryUpdate,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> LotExpiryUpdateResponse:
    audit_context = _build_audit_context(request)
    metadata = audit_context.get("metadata", {})
    if payload.note:
        metadata = {**metadata, "note": payload.note}

    normalized_type = source_type.strip().lower()
    if normalized_type == "production":
        if not _user_has_permission(session, current_user, "production.edit"):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Permiso insuficiente")
        lot = session.get(ProductionLot, source_id)
        if not lot:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote de producción no encontrado")
        before = {"expiry_date": lot.expiry_date}
        lot.expiry_date = payload.expiry_date
        lot.updated_at = datetime.utcnow()
        session.add(lot)
        diff = _diff_fields(before, {"expiry_date": lot.expiry_date}, ["expiry_date"])
        changes = _build_audit_changes(
            "update-expiry",
            "production_lots",
            lot.id,
            diff=diff,
            metadata=metadata,
        )
        _log_audit(session, "production_lots", lot.id, AuditAction.UPDATE, current_user.id, changes, audit_context=audit_context)
        session.commit()
        return LotExpiryUpdateResponse(source_type="production", source_id=lot.id, expiry_date=lot.expiry_date)

    if normalized_type == "purchase":
        if not _user_has_permission(session, current_user, "purchases.edit"):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Permiso insuficiente")
        item = session.get(PurchaseReceiptItem, source_id)
        if not item:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ítem de compra no encontrado")
        before = {"expiry_date": item.expiry_date}
        item.expiry_date = payload.expiry_date
        item.updated_at = datetime.utcnow()
        session.add(item)
        diff = _diff_fields(before, {"expiry_date": item.expiry_date}, ["expiry_date"])
        changes = _build_audit_changes(
            "update-expiry",
            "purchase_receipt_items",
            item.id,
            diff=diff,
            metadata=metadata,
        )
        _log_audit(
            session,
            "purchase_receipt_items",
            item.id,
            AuditAction.UPDATE,
            current_user.id,
            changes,
            audit_context=audit_context,
        )
        session.commit()
        return LotExpiryUpdateResponse(source_type="purchase", source_id=item.id, expiry_date=item.expiry_date)

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de lote inválido")


api_router.include_router(public_router)
api_router.include_router(router)
